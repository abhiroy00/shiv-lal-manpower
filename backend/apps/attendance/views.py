import io
import calendar
import logging
from datetime import date
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
from .models import Attendance, LeaveRequest
from .serializers import AttendanceSerializer, CheckInSerializer, LeaveRequestSerializer
from .services import process_check_in
from apps.employees.models import Employee


LEAVE_LIMITS = {"cl": 12, "sl": 12, "el": 15, "unpaid": None}

logger = logging.getLogger("attendance.checkin")


def _build_register(year, month, site_id=None, district_id=None, search=None):
    """
    Returns (days_in_month, sundays_set, employee_rows_list).
    employee_rows_list items:
        { emp, days: {day_no: code}, present, late, absent, sunday, working_days }
    code: P=present, L=late, R=review, A=absent, S=sunday
    """
    days_in_month = calendar.monthrange(year, month)[1]
    sundays = {
        d for d in range(1, days_in_month + 1)
        if date(year, month, d).weekday() == 6
    }

    emp_qs = Employee.objects.select_related("site__district__state").exclude(status="inactive")
    if site_id:
        emp_qs = emp_qs.filter(site_id=site_id)
    elif district_id:
        emp_qs = emp_qs.filter(site__district_id=district_id)
    if search:
        emp_qs = emp_qs.filter(full_name__icontains=search) | emp_qs.filter(emp_code__icontains=search)

    emp_ids = list(emp_qs.values_list("id", flat=True))

    att_qs = Attendance.objects.filter(
        employee_id__in=emp_ids,
        date__year=year,
        date__month=month,
    ).values("employee_id", "date", "status")

    att_map = {}  # {emp_id: {day: code}}
    for row in att_qs:
        emp_id  = row["employee_id"]
        day     = row["date"].day
        code    = {"present": "P", "late": "L", "review": "R"}.get(row["status"], "P")
        att_map.setdefault(emp_id, {})[day] = code

    rows = []
    for emp in emp_qs:
        days = {}
        present = late = review = 0
        for d in range(1, days_in_month + 1):
            if d in sundays:
                days[d] = "S"
            elif d in att_map.get(emp.id, {}):
                code = att_map[emp.id][d]
                days[d] = code
                if code == "P":   present += 1
                elif code == "L": late    += 1
                elif code == "R": review  += 1
            else:
                days[d] = "A"
        working_days = days_in_month - len(sundays)
        absent = working_days - present - late - review
        rows.append({
            "emp": emp,
            "days": days,
            "present": present,
            "late": late,
            "review": review,
            "absent": absent,
            "working_days": working_days,
        })

    return days_in_month, sundays, rows


class AttendanceViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["employee", "date", "status", "site"]

    def get_queryset(self):
        user = self.request.user
        base_qs = Attendance.objects.select_related("employee", "site", "reviewed_by")
        if user.role in ("admin", "hr", "supervisor"):
            return base_qs.all()
        # Employees only see their own records
        if user.employee_id:
            return base_qs.filter(employee_id=user.employee_id)
        return base_qs.none()

    def _late_threshold(self):
        from django.conf import settings
        from datetime import time
        h, m = getattr(settings, "LATE_THRESHOLD", (9, 30))
        return time(h, m)

    @action(detail=False, methods=["post"], url_path="bulk-fill")
    def bulk_fill(self, request):
        """HR/Admin: fill attendance for a date range for multiple employees."""
        if request.user.role not in ("admin", "hr"):
            return Response({"detail": "Permission denied."}, status=403)

        from datetime import timedelta
        from django.utils.dateparse import parse_date

        from_date  = parse_date(request.data.get("from_date", "") or "")
        to_date    = parse_date(request.data.get("to_date",   "") or "")
        new_status = request.data.get("status")
        emp_ids    = request.data.get("employee_ids") or []
        overwrite  = bool(request.data.get("overwrite", True))

        if not from_date or not to_date:
            return Response({"detail": "from_date and to_date are required (YYYY-MM-DD)."}, status=400)
        if new_status not in ("present", "late", "absent"):
            return Response({"detail": "status must be present, late, or absent."}, status=400)
        if to_date < from_date:
            return Response({"detail": "to_date must be on or after from_date."}, status=400)
        if (to_date - from_date).days > 3650:
            return Response({"detail": "Date range cannot exceed 10 years."}, status=400)

        emp_qs   = Employee.objects.filter(pk__in=emp_ids) if emp_ids else Employee.objects.filter(status="active")
        emp_list = list(emp_qs.select_related("site"))
        if not emp_list:
            return Response({"detail": "No employees found for given scope."}, status=400)

        # All non-Sunday dates in range
        dates, cur = [], from_date
        while cur <= to_date:
            if cur.weekday() != 6:   # 6 = Sunday
                dates.append(cur)
            cur += timedelta(days=1)

        if not dates:
            return Response({"detail": "No working days in the selected range."}, status=400)

        if new_status == "absent":
            deleted, _ = Attendance.objects.filter(
                employee__in=emp_list, date__in=dates
            ).delete()
            return Response({"deleted": deleted, "created": 0, "skipped": 0,
                             "days": len(dates), "employees": len(emp_list)})

        if overwrite:
            Attendance.objects.filter(employee__in=emp_list, date__in=dates).delete()
            records = [
                Attendance(employee=emp, date=d, status=new_status,
                           site=emp.site, geofence_ok=True)
                for emp in emp_list for d in dates
            ]
            Attendance.objects.bulk_create(records, batch_size=500)
            return Response({"created": len(records), "skipped": 0, "deleted": 0,
                             "days": len(dates), "employees": len(emp_list)})
        else:
            existing = set(
                Attendance.objects.filter(employee__in=emp_list, date__in=dates)
                .values_list("employee_id", "date")
            )
            new_records = [
                Attendance(employee=emp, date=d, status=new_status,
                           site=emp.site, geofence_ok=True)
                for emp in emp_list for d in dates
                if (emp.id, d) not in existing
            ]
            Attendance.objects.bulk_create(new_records, batch_size=500)
            return Response({"created": len(new_records),
                             "skipped": len(emp_list) * len(dates) - len(new_records),
                             "deleted": 0, "days": len(dates), "employees": len(emp_list)})

    @action(detail=False, methods=["post"], url_path="bulk-approve")
    def bulk_approve(self, request):
        """HR/Admin: approve all 'review' attendance records in one shot."""
        if request.user.role not in ("admin", "hr"):
            return Response({"detail": "Permission denied."}, status=403)
        threshold = self._late_threshold()
        qs = Attendance.objects.filter(status="review")
        updated = 0
        for att in qs:
            att.status = "late" if (att.check_in_time and att.check_in_time > threshold) else "present"
            att.reviewed_by = request.user
            att.reviewed_at = timezone.now()
            att.review_note = "Bulk approved"
            att.save()
            updated += 1
        return Response({"approved": updated})

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        """HR/Admin: approve a review attendance → auto-determine present/late from check_in_time."""
        if request.user.role not in ("admin", "hr"):
            return Response({"detail": "Permission denied."}, status=403)
        att = self.get_object()
        if att.status != "review":
            return Response({"detail": "Only 'review' attendance can be approved."}, status=400)
        threshold = self._late_threshold()
        att.status      = "late" if (att.check_in_time and att.check_in_time > threshold) else "present"
        att.reviewed_by = request.user
        att.reviewed_at = timezone.now()
        att.review_note = request.data.get("note", "")
        att.save()
        return Response(AttendanceSerializer(att).data)

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        """HR/Admin: reject a review attendance → mark absent."""
        if request.user.role not in ("admin", "hr"):
            return Response({"detail": "Permission denied."}, status=403)
        att = self.get_object()
        if att.status != "review":
            return Response({"detail": "Only 'review' attendance can be rejected."}, status=400)
        att.status      = "absent"
        att.reviewed_by = request.user
        att.reviewed_at = timezone.now()
        att.review_note = request.data.get("note", "")
        att.save()
        return Response(AttendanceSerializer(att).data)

    @action(detail=False, methods=["post"], url_path="delete-selfies")
    def delete_selfies(self, request):
        """Admin: bulk-delete check-in selfie images. Keeps the attendance
        records — only clears/removes the image files.
        Body (one of):
          { "ids": [1, 2, 3] }   → delete selfies for those attendance records
          { "date": "2026-06-28" } → delete all selfies captured on that date
          { "all": true }        → delete EVERY stored selfie (use with care)
        """
        if request.user.role != "admin":
            return Response({"detail": "Permission denied — admin only."}, status=403)

        ids      = request.data.get("ids")
        date_str = request.data.get("date")
        wipe_all = request.data.get("all") is True

        # Only consider records that actually have a selfie file
        qs = Attendance.objects.exclude(selfie="").exclude(selfie__isnull=True)
        if ids:
            qs = qs.filter(id__in=ids)
        elif date_str:
            qs = qs.filter(date=date_str)
        elif not wipe_all:
            return Response(
                {"detail": "Provide 'ids', 'date', or 'all': true to scope the deletion."},
                status=400,
            )

        deleted = 0
        for att in qs:
            if att.selfie:
                att.selfie.delete(save=False)   # remove the file from storage
                att.selfie = None
                att.save(update_fields=["selfie"])
                deleted += 1
        return Response({"deleted": deleted})

    @action(detail=True, methods=["delete"], url_path="selfie")
    def delete_selfie(self, request, pk=None):
        """Admin: delete a single record's selfie image (keeps the record)."""
        if request.user.role != "admin":
            return Response({"detail": "Permission denied — admin only."}, status=403)
        att = self.get_object()
        if not att.selfie:
            return Response({"detail": "No selfie on this record."}, status=400)
        att.selfie.delete(save=False)
        att.selfie = None
        att.save(update_fields=["selfie"])
        return Response({"deleted": 1})

    @action(detail=False, methods=["post"], url_path="mark",
            permission_classes=[IsAuthenticated])
    def mark(self, request):
        """HR: set or clear attendance for a single employee+date.
        Body: { employee_id, date, status: "present"|"late"|"absent" }
        """
        from apps.common.permissions import IsAdminHR
        # Only admin/hr can mark attendance manually
        if request.user.role not in ("admin", "hr"):
            return Response({"detail": "Permission denied."}, status=403)

        emp_id   = request.data.get("employee_id")
        date_str = request.data.get("date")
        new_status = request.data.get("status")

        if not emp_id or not date_str or new_status not in ("present", "late", "absent"):
            return Response({"detail": "employee_id, date and status (present/late/absent) are required."}, status=400)

        if new_status == "absent":
            Attendance.objects.filter(employee_id=emp_id, date=date_str).delete()
            return Response({"detail": "Attendance cleared.", "status": "absent"})

        try:
            emp = Employee.objects.get(pk=emp_id)
        except Employee.DoesNotExist:
            return Response({"detail": "Employee not found."}, status=404)

        att, _ = Attendance.objects.update_or_create(
            employee_id=emp_id,
            date=date_str,
            defaults={
                "status": new_status,
                "site":   emp.site,
                "geofence_ok": True,
            },
        )
        # If no check_in_time yet, set one so the mobile app can display it
        if not att.check_in_time:
            from datetime import datetime, time as dt_time
            from django.utils.dateparse import parse_time
            raw_time = request.data.get("check_in_time")
            if raw_time:
                att.check_in_time = parse_time(raw_time)
            else:
                threshold = self._late_threshold()
                att.check_in_time = datetime.now().time() if new_status == "late" else threshold
            att.save(update_fields=["check_in_time"])
        return Response({"detail": "Attendance saved.", "status": att.status})

    @action(detail=False, methods=["get"], url_path="register")
    def register(self, request):
        today = date.today()
        year  = int(request.query_params.get("year",  today.year))
        month = int(request.query_params.get("month", today.month))
        site_id     = request.query_params.get("site")
        district_id = request.query_params.get("district")
        search      = request.query_params.get("search")

        days_in_month, sundays, rows = _build_register(year, month, site_id, district_id, search)

        employees = []
        for r in rows:
            emp = r["emp"]
            employees.append({
                "id":          emp.id,
                "emp_code":    emp.emp_code,
                "full_name":   emp.full_name,
                "designation": emp.designation,
                "site_name":   emp.site.name if emp.site else "",
                "days":        r["days"],
                "present":     r["present"],
                "late":        r["late"],
                "review":      r["review"],
                "absent":      r["absent"],
                "working_days": r["working_days"],
            })

        return Response({
            "year":          year,
            "month":         month,
            "days_in_month": days_in_month,
            "sundays":       sorted(sundays),
            "employees":     employees,
        })

    @action(detail=False, methods=["get"], url_path="register/export")
    def register_export(self, request):
        today = date.today()
        year  = int(request.query_params.get("year",  today.year))
        month = int(request.query_params.get("month", today.month))
        site_id     = request.query_params.get("site")
        district_id = request.query_params.get("district")

        days_in_month, sundays, rows = _build_register(year, month, site_id, district_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Register"

        # ── Styles ──────────────────────────────────────────────
        thin       = Side(style="thin", color="C0C8D8")
        bdr        = Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr        = Alignment(horizontal="center", vertical="center")
        navy_fill  = PatternFill("solid", fgColor="1E3563")
        navy_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        sun_fill   = PatternFill("solid", fgColor="FBE6E5")
        sun_font   = Font(name="Calibri", color="C0392B", bold=True, size=9)
        p_fill     = PatternFill("solid", fgColor="E1F4EC")
        a_fill     = PatternFill("solid", fgColor="FBE6E5")
        l_fill     = PatternFill("solid", fgColor="FBF1DC")
        r_fill     = PatternFill("solid", fgColor="EDE7F6")
        hdr_fill   = PatternFill("solid", fgColor="F4F6FA")
        hdr_font   = Font(name="Calibri", bold=True, size=9, color="1E3563")
        tot_font   = Font(name="Calibri", bold=True, size=9)
        cell_font  = Font(name="Calibri", size=9)

        month_name = calendar.month_name[month]

        # Title
        total_cols = 5 + days_in_month + 5
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title = ws.cell(row=1, column=1,
            value=f"ATTENDANCE REGISTER – {month_name.upper()} {year}")
        title.font = Font(name="Calibri", bold=True, size=14, color="1E3563")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Column headers
        FIXED_COLS = ["Sr", "Emp Code", "Name", "Designation", "Site"]
        TOTAL_COLS = ["P", "L", "R", "A", "W/D"]
        TOTAL_LABELS = ["Present", "Late", "Review", "Absent", "Working Days"]

        ws.row_dimensions[2].height = 22
        for ci, label in enumerate(FIXED_COLS, start=1):
            c = ws.cell(row=2, column=ci, value=label)
            c.font = navy_font; c.fill = navy_fill; c.alignment = ctr; c.border = bdr

        for d in range(1, days_in_month + 1):
            col = d + len(FIXED_COLS)
            day_date = date(year, month, d)
            label = f"{d}\n{day_date.strftime('%a')[:2]}"
            c = ws.cell(row=2, column=col, value=label)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = bdr
            if d in sundays:
                c.font = sun_font; c.fill = sun_fill
            else:
                c.font = navy_font; c.fill = navy_fill

        for ti, (short, long) in enumerate(zip(TOTAL_COLS, TOTAL_LABELS)):
            col = len(FIXED_COLS) + days_in_month + 1 + ti
            c = ws.cell(row=2, column=col, value=short)
            c.font = navy_font; c.fill = navy_fill; c.alignment = ctr; c.border = bdr
            ws.column_dimensions[get_column_letter(col)].width = 5

        # Column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 11
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 17
        ws.column_dimensions["E"].width = 18
        for d in range(1, days_in_month + 1):
            ws.column_dimensions[get_column_letter(d + len(FIXED_COLS))].width = 3.5

        CODE_FILLS = {"P": p_fill, "A": a_fill, "L": l_fill, "R": r_fill}
        CODE_FONTS = {
            "P": Font(name="Calibri", bold=True, size=8, color="15966A"),
            "A": Font(name="Calibri", bold=True, size=8, color="C0392B"),
            "L": Font(name="Calibri", bold=True, size=8, color="C98A12"),
            "R": Font(name="Calibri", bold=True, size=8, color="7B1FA2"),
            "S": Font(name="Calibri", bold=True, size=8, color="C0392B"),
        }

        for ri, row in enumerate(rows, start=3):
            emp = row["emp"]
            alt = (ri % 2 == 0)
            row_bg = PatternFill("solid", fgColor="F4F6FA") if alt else None

            for ci, val in enumerate([ri - 2, emp.emp_code, emp.full_name,
                                       emp.designation, emp.site.name if emp.site else ""], start=1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = cell_font; c.border = bdr
                if alt and ci > 1: c.fill = row_bg
                if ci == 1: c.alignment = ctr

            for d in range(1, days_in_month + 1):
                code = row["days"].get(d, "A")
                col  = d + len(FIXED_COLS)
                c = ws.cell(row=ri, column=col, value=code)
                c.alignment = ctr; c.border = bdr
                c.font = CODE_FONTS.get(code, cell_font)
                if code in CODE_FILLS:
                    c.fill = CODE_FILLS[code]
                elif code == "S":
                    c.fill = sun_fill

            for ti, key in enumerate(["present", "late", "review", "absent", "working_days"]):
                col = len(FIXED_COLS) + days_in_month + 1 + ti
                c = ws.cell(row=ri, column=col, value=row[key])
                c.font = tot_font; c.alignment = ctr; c.border = bdr
                if alt: c.fill = row_bg

        # Legend
        legend_row = len(rows) + 4
        ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True, size=9)
        for li, (code, desc) in enumerate([("P","Present"), ("L","Late"), ("R","Under Review"), ("A","Absent"), ("S","Sunday")]):
            col = 2 + li
            c = ws.cell(row=legend_row, column=col, value=f"{code}={desc}")
            c.font = CODE_FONTS.get(code, Font(size=9))

        ws.freeze_panes = "F3"

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"attendance_register_{year}_{month:02d}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{fname}"'
        return response


class CheckInView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Diagnostic: log exactly what arrived so we can tell client vs server issues.
        logger.info(
            "[check-in] user=%s content_type=%s FILES=%s data_keys=%s",
            getattr(request.user, "phone", "?"),
            request.content_type,
            {k: (getattr(v, "name", None), getattr(v, "size", None)) for k, v in request.FILES.items()},
            list(request.data.keys()),
        )

        serializer = CheckInSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            employee = request.user.employee
        except Exception:
            return Response({"detail": "No employee linked to this account."}, status=400)

        selfie = serializer.validated_data.get("selfie")
        logger.info("[check-in] user=%s parsed selfie=%s",
                    getattr(request.user, "phone", "?"),
                    getattr(selfie, "name", None) if selfie else None)

        attendance, created = process_check_in(
            employee,
            serializer.validated_data["lat"],
            serializer.validated_data["lng"],
            selfie,
        )
        if not created:
            return Response({"detail": "Already checked in today."}, status=400)
        logger.info("[check-in] user=%s saved attendance id=%s selfie_field=%s",
                    getattr(request.user, "phone", "?"), attendance.id,
                    bool(attendance.selfie))
        return Response(AttendanceSerializer(attendance).data, status=201)


class TodaySummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = date.today()
        total = Employee.objects.filter(status="active").count()
        present = Attendance.objects.filter(date=today, status__in=["present", "late"]).count()
        absent = total - present
        review = Attendance.objects.filter(date=today, status="review").count()
        return Response({
            "date": today,
            "total_active": total,
            "present": present,
            "absent": absent,
            "under_review": review,
        })


class MyTodayView(APIView):
    """Employee: today's check-in status."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            emp = request.user.employee
        except Exception:
            return Response({"detail": "No employee linked."}, status=400)
        today = date.today()
        try:
            att = Attendance.objects.get(employee=emp, date=today)
            return Response({
                "checked_in":      True,
                "check_in_time":   att.check_in_time,
                "check_out_time":  att.check_out_time,
                "status":          att.status,
                "geofence_ok":     att.geofence_ok,
            })
        except Attendance.DoesNotExist:
            return Response({"checked_in": False, "check_in_time": None,
                             "check_out_time": None, "status": None})


class CheckOutView(APIView):
    """Employee: mark check-out for today."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            emp = request.user.employee
        except Exception:
            return Response({"detail": "No employee linked."}, status=400)
        today = date.today()
        try:
            att = Attendance.objects.get(employee=emp, date=today)
        except Attendance.DoesNotExist:
            return Response({"detail": "No check-in found for today."}, status=400)
        if att.check_out_time:
            return Response({"detail": "Already checked out."}, status=400)
        att.check_out_time = timezone.localtime(timezone.now()).time()
        att.save()
        return Response(AttendanceSerializer(att).data)


class MyAttendanceView(APIView):
    """Employee: own attendance for a given month/year."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            emp = request.user.employee
        except Exception:
            return Response({"detail": "No employee linked."}, status=400)
        today  = date.today()
        year   = int(request.query_params.get("year",  today.year))
        month  = int(request.query_params.get("month", today.month))
        import calendar as cal
        days_in_month = cal.monthrange(year, month)[1]
        sundays = {d for d in range(1, days_in_month + 1) if date(year, month, d).weekday() == 6}

        atts = Attendance.objects.filter(employee=emp, date__year=year, date__month=month).order_by("date")
        att_map = {a.date.day: a for a in atts}

        days = []
        present = late = absent = leave_days = 0
        for d in range(1, days_in_month + 1):
            if d in sundays:
                days.append({"day": d, "code": "S", "status": "sunday"})
            elif d in att_map:
                a = att_map[d]
                code = {"present": "P", "late": "L", "review": "R"}.get(a.status, "P")
                if code == "P": present += 1
                elif code == "L": late += 1
                days.append({
                    "day": d, "code": code, "status": a.status,
                    "check_in_time":  str(a.check_in_time)  if a.check_in_time  else None,
                    "check_out_time": str(a.check_out_time) if a.check_out_time else None,
                })
            elif date(year, month, d) > today:
                days.append({"day": d, "code": "", "status": "future"})
            else:
                on_leave = LeaveRequest.objects.filter(
                    employee=emp, status="approved",
                    from_date__lte=date(year, month, d),
                    to_date__gte=date(year, month, d),
                ).exists()
                if on_leave:
                    days.append({"day": d, "code": "LE", "status": "leave"})
                    leave_days += 1
                else:
                    days.append({"day": d, "code": "A", "status": "absent"})
                    absent += 1
        working_days = days_in_month - len(sundays)
        return Response({
            "year": year, "month": month, "days_in_month": days_in_month,
            "days": days,
            "summary": {"present": present, "late": late, "absent": absent,
                        "leave": leave_days, "working_days": working_days},
        })


class AttendanceMapView(APIView):
    """Return check-in locations for a given date (for map display)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        target_date = request.query_params.get("date", str(date.today()))
        site_id     = request.query_params.get("site")
        qs = Attendance.objects.select_related("employee__site").filter(
            date=target_date,
            lat__isnull=False,
            lng__isnull=False,
        )
        if site_id:
            qs = qs.filter(employee__site_id=site_id)
        data = []
        for a in qs:
            data.append({
                "id":           a.id,
                "emp_code":     a.employee.emp_code,
                "full_name":    a.employee.full_name,
                "designation":  a.employee.designation,
                "site_name":    a.employee.site.name if a.employee.site else None,
                "lat":          float(a.lat),
                "lng":          float(a.lng),
                "check_in_time": str(a.check_in_time) if a.check_in_time else None,
                "check_out_time":str(a.check_out_time) if a.check_out_time else None,
                "status":       a.status,
                "geofence_ok":  a.geofence_ok,
            })
        return Response({"date": target_date, "count": len(data), "records": data})


class LeaveRequestViewSet(viewsets.ModelViewSet):
    serializer_class   = LeaveRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role in ("admin", "hr"):
            return LeaveRequest.objects.select_related("employee", "reviewed_by").all()
        try:
            return LeaveRequest.objects.filter(employee=user.employee)
        except Exception:
            return LeaveRequest.objects.none()

    def perform_create(self, serializer):
        try:
            emp = self.request.user.employee
        except Exception:
            from rest_framework.exceptions import ValidationError
            raise ValidationError("No employee linked to your account.")
        serializer.save(employee=emp)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        leave = self.get_object()
        leave.status      = "approved"
        leave.reviewed_by = request.user
        leave.reviewed_at = timezone.now()
        leave.review_note = request.data.get("note", "")
        leave.save()
        return Response(LeaveRequestSerializer(leave).data)

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        leave = self.get_object()
        leave.status      = "rejected"
        leave.reviewed_by = request.user
        leave.reviewed_at = timezone.now()
        leave.review_note = request.data.get("note", "")
        leave.save()
        return Response(LeaveRequestSerializer(leave).data)

    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        leave = self.get_object()
        try:
            emp = request.user.employee
        except Exception:
            return Response({"detail": "No employee linked."}, status=400)
        if leave.employee != emp:
            return Response({"detail": "You can only cancel your own leave requests."}, status=403)
        if leave.status != "pending":
            return Response({"detail": "Only pending leave requests can be cancelled."}, status=400)
        leave.delete()
        return Response({"detail": "Leave request cancelled."}, status=200)


class LeaveBalanceView(APIView):
    """Employee: leave usage summary for the current year."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            emp = request.user.employee
        except Exception:
            return Response({"detail": "No employee linked."}, status=400)

        current_year = date.today().year
        leaves = LeaveRequest.objects.filter(
            employee=emp,
            from_date__year=current_year,
        )

        balance = {}
        for lt, limit in LEAVE_LIMITS.items():
            used = sum(l.days for l in leaves if l.leave_type == lt and l.status == "approved")
            pending = sum(l.days for l in leaves if l.leave_type == lt and l.status == "pending")
            balance[lt] = {
                "used":      used,
                "pending":   pending,
                "limit":     limit,
                "remaining": max(limit - used, 0) if limit is not None else None,
            }

        return Response({"year": current_year, "balance": balance})
