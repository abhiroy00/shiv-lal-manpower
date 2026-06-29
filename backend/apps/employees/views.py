import io
import logging
import traceback
from datetime import date
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Employee, EmployeeDocument
from .serializers import EmployeeSerializer, EmployeeListSerializer, EmployeeDocumentSerializer
from .filters import EmployeeFilter
from apps.common.permissions import IsAdminHR


logger = logging.getLogger("employees.import")


class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.select_related("site__district__state", "user_account").all()
    permission_classes = [IsAuthenticated]
    filterset_class = EmployeeFilter
    search_fields = ["emp_code", "full_name", "phone", "designation"]
    ordering_fields = ["full_name", "date_joined", "created_at"]
    ordering = ["-created_at"]

    def get_serializer_class(self):
        if self.action == "list":
            return EmployeeListSerializer
        return EmployeeSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        employee = serializer.save()

        from apps.accounts.models import User
        credentials = None
        if not User.objects.filter(phone=employee.phone).exists():
            User.objects.create_user(
                phone=employee.phone,
                password=employee.phone,
                full_name=employee.full_name,
                role="employee",
                employee=employee,
            )
            credentials = {
                "phone": employee.phone,
                "default_password": employee.phone,
            }

        data = EmployeeSerializer(employee, context={"request": request}).data
        if credentials:
            data["credentials"] = credentials
        return Response(data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"], url_path="reset-password",
            permission_classes=[IsAdminHR])
    def reset_password(self, request, pk=None):
        employee = self.get_object()
        from apps.accounts.models import User
        try:
            user = employee.user_account
        except Exception:
            return Response({"detail": "No login account found for this employee."}, status=400)
        user.set_password(employee.phone)
        user.password_changed_at = None
        user.save()
        return Response({
            "detail": "Password reset successfully.",
            "phone": employee.phone,
            "default_password": employee.phone,
        })

    @action(detail=True, methods=["post"], url_path="create-login",
            permission_classes=[IsAdminHR])
    def create_login(self, request, pk=None):
        """Create (or re-link) a mobile app login for an employee who doesn't have one."""
        employee = self.get_object()
        from apps.accounts.models import User
        try:
            if employee.user_account is not None:
                return Response({"detail": "This employee already has a login. Use Reset Password instead."}, status=400)
        except Exception:
            pass  # No account exists — proceed

        existing = User.objects.filter(phone=employee.phone).first()
        if existing:
            # User with this phone exists but employee FK is NULL (e.g. old account after re-import)
            if existing.employee_id is not None and existing.employee_id != employee.id:
                return Response(
                    {"detail": f"Phone {employee.phone} is already linked to a different employee account."},
                    status=400,
                )
            existing.employee = employee
            existing.role = "employee"
            existing.set_password(employee.phone)
            existing.password_changed_at = None
            existing.save()
        else:
            User.objects.create_user(
                phone=employee.phone,
                password=employee.phone,
                full_name=employee.full_name,
                role="employee",
                employee=employee,
            )
        return Response({
            "detail": "Login created successfully.",
            "phone": employee.phone,
            "default_password": employee.phone,
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"], url_path="transfer")
    def transfer(self, request, pk=None):
        """Move employee to a different site. Body: { site_id }"""
        employee = self.get_object()
        site_id  = request.data.get("site_id")
        from apps.deployment.models import Site
        if site_id:
            try:
                site = Site.objects.get(pk=site_id)
            except Site.DoesNotExist:
                return Response({"detail": "Site not found."}, status=400)
            employee.site = site
        else:
            employee.site = None
        employee.save()
        return Response(EmployeeSerializer(employee).data)

    @action(detail=False, methods=["post"], url_path="bulk-transfer")
    def bulk_transfer(self, request):
        """Move multiple employees to a site. Body: { employee_ids: [], site_id }"""
        emp_ids = request.data.get("employee_ids", [])
        site_id = request.data.get("site_id")
        from apps.deployment.models import Site
        try:
            site = Site.objects.get(pk=site_id) if site_id else None
        except Site.DoesNotExist:
            return Response({"detail": "Site not found."}, status=400)
        updated = Employee.objects.filter(pk__in=emp_ids).update(site=site)
        return Response({"transferred": updated, "site": site.name if site else None})

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """Return a blank Excel template that matches the import format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        hdr_fill = PatternFill("solid", fgColor="1E3563")
        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D0D7E5")
        bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

        HEADERS = [
            ("Emp Code", 14), ("Full Name", 22), ("Designation", 20), ("Phone", 14),
            ("State", 18), ("District", 18), ("Site", 24), ("Date Joined", 14), ("Status", 12),
            ("Date of Birth", 14), ("Address", 30),
            ("UAN", 16), ("ESIC No", 16), ("Aadhar", 15), ("PAN", 13),
            ("Bank Account", 20), ("IFSC", 13), ("TDS", 14),
            ("Basic", 12), ("HRA", 10), ("DA", 10), ("Other Allowances", 16),
        ]
        ws.row_dimensions[1].height = 28
        for col, (label, width) in enumerate(HEADERS, 1):
            c = ws.cell(1, col, label)
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = bdr
            ws.column_dimensions[get_column_letter(col)].width = width

        # Sample row
        sample = ["", "John Doe", "Security Guard", "9876543210", "Jharkhand", "Deoghar",
                  "DEOGHAR COLLEGE DEOGHAR", "2024-01-15", "active", "1995-06-20", "",
                  "", "", "", "", "", "", "",
                  "10365", "4146", "0", "0"]
        for col, val in enumerate(sample, 1):
            c = ws.cell(2, col, val); c.border = bdr

        # Instructions sheet
        ws2 = wb.create_sheet("Instructions")
        ws2["A1"] = "Import Instructions – Shiv Lal Manpower Portal"
        ws2["A1"].font = Font(bold=True, size=13, color="1E3563")
        notes = [
            ["", ""],
            ["Column", "Notes"],
            ["Emp Code", "Optional — leave blank to auto-assign"],
            ["Full Name", "REQUIRED"],
            ["Designation", "REQUIRED"],
            ["Phone", "REQUIRED — 10-digit, must be unique"],
            ["State", "Optional — State name (e.g. Jharkhand). Used to narrow site search."],
            ["District", "Optional — District name (e.g. Deoghar). Used to narrow site search."],
            ["Site", "Optional — Site/Office name. Best matched with State + District."],
            ["Date Joined", "REQUIRED — format: YYYY-MM-DD (e.g. 2024-01-15)"],
            ["Status", "Optional — active / on_leave / inactive (default: active)"],
            ["Date of Birth", "Optional — format: YYYY-MM-DD (e.g. 1995-06-20)"],
            ["Address", "Optional — residential address"],
            ["UAN / ESIC No / Aadhar / PAN / Bank Account / IFSC / TDS", "All optional"],
            ["Basic", "Optional — monthly basic salary (e.g. 10365). Creates/updates salary structure."],
            ["HRA", "Optional — House Rent Allowance (e.g. 4146)"],
            ["DA", "Optional — Dearness Allowance (e.g. 0)"],
            ["Other Allowances", "Optional — any other allowance (e.g. 0)"],
            ["", ""],
            ["IMPORTANT — numeric fields", "Format Aadhar, UAN, ESIC No, Bank Account cells as TEXT in Excel before entering numbers. "
             "Otherwise Excel stores them as floating-point (e.g. 290909953973.0) which adds a decimal and may fail validation."],
            ["", ""],
            ["NOTE", "Delete the sample row (row 2) before importing your data."],
        ]
        for row in notes:
            ws2.append(row)
        ws2.column_dimensions["A"].width = 45
        ws2.column_dimensions["B"].width = 50

        ws.freeze_panes = "A2"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        response = HttpResponse(buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="employee_import_template.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path="import")
    def import_employees(self, request):
        """Bulk-create employees from an uploaded Excel file."""
        import time, uuid
        imp_id = uuid.uuid4().hex[:8]          # short id to correlate log lines
        t_start = time.time()

        if request.user.role not in ("admin", "hr"):
            logger.warning("[import %s] DENIED — user=%s role=%s", imp_id,
                           getattr(request.user, "phone", "?"), getattr(request.user, "role", "?"))
            return Response({"detail": "Permission denied."}, status=403)
        file = request.FILES.get("file")
        if not file:
            logger.warning("[import %s] no file provided by user=%s", imp_id,
                           getattr(request.user, "phone", "?"))
            return Response({"detail": "No file provided. Upload an .xlsx file."}, status=400)

        logger.info("[import %s] START — user=%s file=%r size=%s bytes", imp_id,
                    getattr(request.user, "phone", "?"), getattr(file, "name", "?"),
                    getattr(file, "size", "?"))

        from openpyxl import load_workbook
        from django.db import IntegrityError
        from apps.deployment.models import Site
        from apps.accounts.models import User
        import re

        try:
            wb = load_workbook(file, data_only=True)
        except Exception:
            logger.error("[import %s] FAILED to open workbook:\n%s", imp_id, traceback.format_exc())
            return Response({"detail": "Invalid Excel file. Please use the provided template."}, status=400)

        ws = wb.active
        if ws.max_row < 2:
            logger.warning("[import %s] empty file — max_row=%s", imp_id, ws.max_row)
            return Response({"detail": "File is empty — no data rows found."}, status=400)

        # Map header name → 0-based index
        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        col = {h: i for i, h in enumerate(headers)}

        # Log which columns were detected vs which the importer looks for, so a
        # header typo (e.g. "Aadhaar" vs "Aadhar") is obvious in the logs.
        EXPECTED_COLS = ["Emp Code", "Full Name", "Designation", "Phone", "State",
                         "District", "Site", "Date Joined", "Status",
                         "Date of Birth", "Address",
                         "UAN", "ESIC No", "Aadhar", "PAN", "Bank Account", "IFSC", "TDS",
                         "Basic", "HRA", "DA", "Other Allowances"]
        missing_cols = [c for c in EXPECTED_COLS if c not in col]
        logger.info("[import %s] rows=%s headers=%s", imp_id, ws.max_row - 1, headers)
        if missing_cols:
            logger.warning("[import %s] columns NOT found in sheet (will be treated as blank): %s",
                           imp_id, missing_cols)

        def cell_str(row, name):
            idx = col.get(name)
            if idx is None:
                return ""
            v = ws.cell(row, idx + 1).value
            if v is None:
                return ""
            # Excel stores every numeric cell as a Python float (e.g. 290909953973.0).
            # str(float) gives "290909953973.0" which is 14 chars — breaks the 12-char
            # Aadhar length check and silently skips the whole row.  Convert whole-number
            # floats to int first so we get the clean string "290909953973".
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v).strip()

        def cell_date(row, name):
            idx = col.get(name)
            if idx is None:
                return None
            v = ws.cell(row, idx + 1).value
            if v is None:
                return None
            if hasattr(v, "date"):   # datetime from Excel
                return v.date()
            if hasattr(v, "year"):   # date object
                return v
            from datetime import datetime
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y"):
                try:
                    return datetime.strptime(str(v).strip(), fmt).date()
                except ValueError:
                    pass
            return None

        STATUS_MAP = {
            "active": "active", "on_leave": "on_leave", "inactive": "inactive",
            "Active": "active", "On Leave": "on_leave", "Inactive": "inactive",
        }

        # Find next auto emp_code number
        existing_codes = list(Employee.objects.values_list("emp_code", flat=True))
        max_num = 0
        for code in existing_codes:
            m = re.search(r"\d+", str(code))
            if m:
                max_num = max(max_num, int(m.group()))

        created = 0
        updated = 0
        skipped = 0
        errors = []
        warnings = []

        for row_idx in range(2, ws.max_row + 1):
            full_name = cell_str(row_idx, "Full Name")
            if not full_name:
                continue   # blank row

            phone       = cell_str(row_idx, "Phone")
            designation = cell_str(row_idx, "Designation")
            date_joined = cell_date(row_idx, "Date Joined")

            row_errors = []
            if not phone:        row_errors.append("Phone is required")
            if not designation:  row_errors.append("Designation is required")
            if not date_joined:  row_errors.append("Date Joined is required/invalid (use YYYY-MM-DD)")
            if row_errors:
                errors.append({"row": row_idx, "name": full_name, "error": "; ".join(row_errors)})
                skipped += 1
                logger.info("[import %s] row %s SKIP (%s) — %s", imp_id, row_idx,
                            full_name, "; ".join(row_errors))
                continue

            emp_code = cell_str(row_idx, "Emp Code")
            if not emp_code:
                max_num += 1
                emp_code = str(max_num).zfill(4)

            site_name    = cell_str(row_idx, "Site")
            state_name   = cell_str(row_idx, "State")
            district_name = cell_str(row_idx, "District")
            site = None
            if site_name:
                qs_site = Site.objects.select_related("district__state")
                if state_name and district_name:
                    site = qs_site.filter(
                        name__iexact=site_name,
                        district__name__iexact=district_name,
                        district__state__name__iexact=state_name,
                    ).first()
                elif district_name:
                    site = qs_site.filter(
                        name__iexact=site_name,
                        district__name__iexact=district_name,
                    ).first()
                elif state_name:
                    site = qs_site.filter(
                        name__iexact=site_name,
                        district__state__name__iexact=state_name,
                    ).first()
                if site is None:
                    site = qs_site.filter(name__iexact=site_name).first()
                if site is None:
                    logger.warning("[import %s] row %s (%s) — site not matched: site=%r district=%r state=%r",
                                   imp_id, row_idx, full_name, site_name, district_name, state_name)

            status_val = STATUS_MAP.get(cell_str(row_idx, "Status"), "active")

            # Read employee profile fields
            dob_val     = cell_date(row_idx, "Date of Birth")
            address_val = cell_str(row_idx, "Address")

            # Read compliance / banking fields
            uan_val    = cell_str(row_idx, "UAN")
            esic_val   = cell_str(row_idx, "ESIC No")
            aadhar_val = cell_str(row_idx, "Aadhar")
            pan_val    = cell_str(row_idx, "PAN")
            bank_val   = cell_str(row_idx, "Bank Account")
            ifsc_val   = cell_str(row_idx, "IFSC")
            tds_val    = cell_str(row_idx, "TDS")

            # Read salary fields
            def _decimal(name):
                from decimal import Decimal, InvalidOperation
                v = cell_str(row_idx, name)
                if not v:
                    return None
                try:
                    return Decimal(v).quantize(Decimal("0.01"))
                except InvalidOperation:
                    return None

            basic_val  = _decimal("Basic")
            hra_val    = _decimal("HRA")
            da_val     = _decimal("DA")
            other_val  = _decimal("Other Allowances")

            # Critical identity fields — skip the whole row if these are invalid
            if len(phone) > 15:
                errors.append({"row": row_idx, "name": full_name,
                                "error": f"Phone too long ({len(phone)} chars, max 15)"})
                skipped += 1
                logger.info("[import %s] row %s SKIP (%s) — phone too long: %r",
                            imp_id, row_idx, full_name, phone)
                continue
            if len(emp_code) > 20:
                errors.append({"row": row_idx, "name": full_name,
                                "error": f"Emp code too long ({len(emp_code)} chars, max 20)"})
                skipped += 1
                logger.info("[import %s] row %s SKIP (%s) — emp_code too long: %r",
                            imp_id, row_idx, full_name, emp_code)
                continue

            # Compliance fields — blank out any field that's too long, still import the row
            # (common cause: IFSC typo with 12 chars, Aadhar as float "290909953973.0" → fixed
            # by cell_str, but kept here as a safety net for bad data)
            row_warns = []
            if len(uan_val) > 20:
                row_warns.append(f"UAN cleared ({len(uan_val)} chars > max 20)"); uan_val = ""
            if len(esic_val) > 20:
                row_warns.append(f"ESIC No cleared ({len(esic_val)} chars > max 20)"); esic_val = ""
            if len(aadhar_val) > 12:
                row_warns.append(f"Aadhar cleared ({len(aadhar_val)} chars > max 12)"); aadhar_val = ""
            if len(pan_val) > 10:
                row_warns.append(f"PAN cleared ({len(pan_val)} chars > max 10)"); pan_val = ""
            if len(bank_val) > 20:
                row_warns.append(f"Bank Account cleared ({len(bank_val)} chars > max 20)"); bank_val = ""
            if len(ifsc_val) > 11:
                row_warns.append(f"IFSC cleared ({len(ifsc_val)} chars > max 11)"); ifsc_val = ""
            if len(tds_val) > 30:
                row_warns.append(f"TDS cleared ({len(tds_val)} chars > max 30)"); tds_val = ""

            if row_warns:
                logger.warning("[import %s] row %s (%s) — field(s) cleared: %s",
                               imp_id, row_idx, full_name, "; ".join(row_warns))

            # Per-row parsed values (DEBUG level — enable verbose logging to see these)
            logger.debug("[import %s] row %s parsed: emp_code=%r phone=%r uan=%r esic=%r "
                         "aadhar=%r pan=%r bank=%r ifsc=%r tds=%r site=%s",
                         imp_id, row_idx, emp_code, phone, uan_val, esic_val, aadhar_val,
                         pan_val, bank_val, ifsc_val, tds_val, getattr(site, "id", None))

            # -- UPSERT: find existing employee by emp_code or phone --
            emp = None
            if emp_code:
                emp = Employee.objects.filter(emp_code__iexact=emp_code).first()
            if emp is None:
                emp = Employee.objects.filter(phone=phone).first()

            try:
                if emp is not None:
                    # PATCH — update all non-empty fields from the sheet
                    update_fields = ["full_name", "designation", "date_joined", "status"]
                    emp.full_name   = full_name
                    emp.designation = designation
                    emp.date_joined = date_joined
                    emp.status      = status_val
                    if site is not None:
                        emp.site = site; update_fields.append("site")
                    if dob_val is not None:
                        emp.date_of_birth = dob_val; update_fields.append("date_of_birth")
                    if address_val:
                        emp.address = address_val;  update_fields.append("address")
                    # Only overwrite compliance fields if the sheet provides a non-empty value
                    if uan_val:    emp.uan          = uan_val;    update_fields.append("uan")
                    if esic_val:   emp.esic_no      = esic_val;   update_fields.append("esic_no")
                    if aadhar_val: emp.aadhar       = aadhar_val; update_fields.append("aadhar")
                    if pan_val:    emp.pan          = pan_val;    update_fields.append("pan")
                    if bank_val:   emp.bank_account = bank_val;   update_fields.append("bank_account")
                    if ifsc_val:   emp.ifsc         = ifsc_val;   update_fields.append("ifsc")
                    if tds_val:    emp.tds          = tds_val;    update_fields.append("tds")
                    emp.save(update_fields=update_fields)
                    updated += 1
                    logger.info("[import %s] row %s UPDATED id=%s (%s) — fields=%s",
                                imp_id, row_idx, emp.id, full_name,
                                [f for f in update_fields if f not in
                                 ("full_name", "designation", "date_joined", "status")] or "core only")
                    if row_warns:
                        warnings.append({"row": row_idx, "name": full_name, "warnings": row_warns})
                    # Ensure the user account is linked
                    existing_user = User.objects.filter(phone=emp.phone).first()
                    if existing_user and not existing_user.employee_id:
                        existing_user.employee = emp
                        existing_user.role = "employee"
                        existing_user.save(update_fields=["employee", "role"])
                else:
                    # CREATE new employee
                    emp = Employee.objects.create(
                        emp_code=emp_code,
                        full_name=full_name,
                        phone=phone,
                        designation=designation,
                        site=site,
                        date_joined=date_joined,
                        status=status_val,
                        date_of_birth=dob_val,
                        address=address_val,
                        uan=uan_val,
                        esic_no=esic_val,
                        aadhar=aadhar_val,
                        pan=pan_val,
                        bank_account=bank_val,
                        ifsc=ifsc_val,
                        tds=tds_val,
                    )
                    existing_user = User.objects.filter(phone=phone).first()
                    if existing_user:
                        if not existing_user.employee_id:
                            existing_user.employee = emp
                            existing_user.role = "employee"
                            existing_user.save(update_fields=["employee", "role"])
                    else:
                        User.objects.create_user(
                            phone=phone, password=phone,
                            full_name=full_name, role="employee", employee=emp,
                        )
                    created += 1
                    logger.info("[import %s] row %s CREATED id=%s (%s) — emp_code=%s phone=%s",
                                imp_id, row_idx, emp.id, full_name, emp_code, phone)
                    if row_warns:
                        warnings.append({"row": row_idx, "name": full_name, "warnings": row_warns})

                # Upsert salary structure if any salary field was provided
                if emp and basic_val is not None:
                    from apps.payroll.models import SalaryStructure
                    from decimal import Decimal as D
                    SalaryStructure.objects.update_or_create(
                        employee=emp,
                        defaults={
                            "basic":            basic_val,
                            "hra":              hra_val   if hra_val   is not None else D("0"),
                            "da":               da_val    if da_val    is not None else D("0"),
                            "other_allowances": other_val if other_val is not None else D("0"),
                        },
                    )
                    logger.info("[import %s] row %s salary — basic=%s hra=%s da=%s other=%s",
                                imp_id, row_idx, basic_val, hra_val, da_val, other_val)

            except Exception as e:
                errors.append({"row": row_idx, "name": full_name, "error": str(e)})
                skipped += 1
                logger.error("[import %s] row %s ERROR (%s) — %s\n%s",
                             imp_id, row_idx, full_name, e, traceback.format_exc())

        elapsed = time.time() - t_start
        logger.info("[import %s] DONE in %.2fs — created=%s updated=%s skipped=%s warnings=%s",
                    imp_id, elapsed, created, updated, skipped, len(warnings))
        return Response({"created": created, "updated": updated, "skipped": skipped,
                         "errors": errors, "warnings": warnings})

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        qs = self.filter_queryset(
            Employee.objects.select_related(
                "site__district__state", "salary_structure"
            ).all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        # ── Styles ──────────────────────────────────────────────
        hdr_fill  = PatternFill("solid", fgColor="1E3563")
        hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin      = Side(style="thin", color="D0D7E5")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill  = PatternFill("solid", fgColor="F4F6FA")
        date_align = Alignment(horizontal="center")

        HEADERS = [
            ("Emp Code",        14),
            ("Full Name",       22),
            ("Designation",     20),
            ("Phone",           14),
            ("Date of Birth",   14),
            ("Address",         30),
            ("State",           14),
            ("District",        18),
            ("Site",            24),
            ("Office Name",     24),
            ("Date Joined",     14),
            ("Status",          12),
            ("UAN",             16),
            ("ESIC No",         16),
            ("Aadhar",          15),
            ("PAN",             13),
            ("Bank Account",    20),
            ("IFSC",            13),
            ("Basic (₹)",      13),
            ("HRA (₹)",        13),
            ("DA (₹)",         13),
            ("Other Allow (₹)", 16),
        ]

        # Header row
        ws.row_dimensions[1].height = 30
        for col, (label, width) in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align
            cell.border = cell_border
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data rows
        STATUS_MAP = {"active": "Active", "on_leave": "On Leave", "inactive": "Inactive"}
        salary_cols = {len(HEADERS) - 3, len(HEADERS) - 2, len(HEADERS) - 1, len(HEADERS)}

        for row_idx, emp in enumerate(qs, start=2):
            is_alt = (row_idx % 2 == 0)
            fill   = alt_fill if is_alt else None

            sal = getattr(emp, "salary_structure", None)
            row_data = [
                emp.emp_code,
                emp.full_name,
                emp.designation,
                emp.phone,
                emp.date_of_birth,
                emp.address,
                emp.site.district.state.name if emp.site and emp.site.district and emp.site.district.state else "",
                emp.site.district.name  if emp.site and emp.site.district else "",
                emp.site.name           if emp.site else "",
                emp.site.office_name    if emp.site else "",
                emp.date_joined,
                STATUS_MAP.get(emp.status, emp.status),
                emp.uan,
                emp.esic_no,
                emp.aadhar,
                emp.pan,
                emp.bank_account,
                emp.ifsc,
                float(sal.basic)            if sal else "",
                float(sal.hra)              if sal else "",
                float(sal.da)               if sal else "",
                float(sal.other_allowances) if sal else "",
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = cell_border
                if fill:
                    cell.fill = fill
                if isinstance(value, date):
                    cell.number_format = "DD-MMM-YYYY"
                    cell.alignment = date_align
                elif col in salary_cols and isinstance(value, float):
                    cell.number_format = '#,##0.00'

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Shiv Lal Manpower – Employee Export"
        ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="1E3563")
        ws2["A3"] = f"Generated on: {date.today().strftime('%d %b %Y')}"
        ws2["A4"] = f"Total records: {qs.count()}"
        ws2["A5"] = f"Filters applied: {dict(request.query_params) or 'None'}"

        # Stream response
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"employees_{date.today().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=["get", "post"])
    def documents(self, request, pk=None):
        employee = self.get_object()
        ctx = {"request": request}
        if request.method == "POST":
            serializer = EmployeeDocumentSerializer(data=request.data, context=ctx)
            serializer.is_valid(raise_exception=True)
            serializer.save(employee=employee)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        docs = employee.documents.all()
        return Response(EmployeeDocumentSerializer(docs, many=True, context=ctx).data)

    @action(detail=True, methods=["delete"], url_path=r"documents/(?P<doc_id>[^/.]+)")
    def delete_document(self, request, pk=None, doc_id=None):
        employee = self.get_object()
        doc = get_object_or_404(EmployeeDocument, pk=doc_id, employee=employee)
        doc.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["get", "post"], url_path="my-documents")
    def my_documents(self, request):
        """Employee self-service: list or upload own documents."""
        if not request.user.employee_id:
            return Response({"detail": "No employee profile linked to this account."}, status=400)
        employee = request.user.employee
        ctx = {"request": request}
        if request.method == "POST":
            f = request.FILES.get("file")
            if f and f.size > 25 * 1024 * 1024:
                return Response({"detail": "File too large. Maximum size is 25 MB."}, status=400)
            serializer = EmployeeDocumentSerializer(data=request.data, context=ctx)
            serializer.is_valid(raise_exception=True)
            serializer.save(employee=employee)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        docs = employee.documents.all().order_by("-uploaded_at")
        return Response(EmployeeDocumentSerializer(docs, many=True, context=ctx).data)

    @action(detail=False, methods=["delete"], url_path=r"my-documents/(?P<doc_id>[^/.]+)")
    def delete_my_document(self, request, doc_id=None):
        """Employee self-service: delete own document."""
        if not request.user.employee_id:
            return Response({"detail": "No employee profile linked to this account."}, status=400)
        employee = request.user.employee
        doc = get_object_or_404(EmployeeDocument, pk=doc_id, employee=employee)
        doc.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
