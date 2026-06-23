import io
import calendar
from decimal import Decimal
from datetime import date
from django.http import HttpResponse
from django.db.models import Sum
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from apps.payroll.models import PayrollRun, Payslip
from apps.common.permissions import IsAdminHR
from .models import PFContribution, ESIContribution, ChallanRun
from .serializers import PFContributionSerializer, ESIContributionSerializer, ChallanRunSerializer
from .services import (
    generate_compliance,
    PF_EMP_RATE, EPS_RATE, EPF_EMP_RATE, EPS_CAP, EDLI_RATE,
    ESI_EMP_RATE, ESI_EMPLR_RATE, ESI_THRESHOLD,
)

# ── Excel style helpers ───────────────────────────────────────
def _tbl_style(ws, max_row, max_col, hdr_fill_hex):
    thin   = Side(style="thin", color="D0D7E5")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy   = PatternFill("solid", fgColor=hdr_fill_hex)
    nfont  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    cfont  = Font(name="Calibri", size=10)
    altbg  = PatternFill("solid", fgColor="F4F6FA")

    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=max_col):
        for c in row:
            c.font = nfont; c.fill = navy; c.border = bdr
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri in range(4, max_row + 1):
        alt = (ri % 2 == 0)
        for c in ws[ri]:
            c.font = cfont; c.border = bdr
            if alt: c.fill = altbg

    ws.freeze_panes = "A4"
    ws.row_dimensions[3].height = 22


# ── Summary endpoint ──────────────────────────────────────────
class ComplianceSummaryView(APIView):
    permission_classes = [IsAdminHR]

    def get(self, request):
        run_id = request.query_params.get("payroll_run")
        if not run_id:
            # Return all challans summary
            challans = ChallanRun.objects.select_related("payroll_run").order_by(
                "-payroll_run__year", "-payroll_run__month"
            )
            data = []
            for c in challans:
                data.append({
                    "id":          c.id,
                    "challan_type": c.challan_type,
                    "month":       c.payroll_run.month,
                    "year":        c.payroll_run.year,
                    "month_label": f"{calendar.month_name[c.payroll_run.month]} {c.payroll_run.year}",
                    "total_amount": float(c.total_amount),
                    "is_filed":    c.is_filed,
                    "filed_on":    c.filed_on,
                    "payroll_run_id": c.payroll_run_id,
                })
            return Response(data)

        # Summary for a specific run
        try:
            run = PayrollRun.objects.get(pk=run_id)
        except PayrollRun.DoesNotExist:
            return Response({"detail": "PayrollRun not found."}, status=404)

        payslips = run.payslips.all()
        basic_total = float(payslips.aggregate(s=Sum("basic"))["s"] or 0)
        gross_total = float(
            payslips.aggregate(
                s=Sum("basic") + Sum("hra") + Sum("da") + Sum("other_allowances")
            )["s"] or 0
        )
        esi_eligible = payslips.filter(
            esi_employee__gt=0
        ).count()

        # PF figures
        pf_emp   = round(basic_total * float(PF_EMP_RATE), 2)
        eps      = min(round(basic_total * float(EPS_RATE), 2),
                       float(EPS_CAP) * payslips.count())
        epf_empl = round(basic_total * float(EPF_EMP_RATE), 2)
        edli     = round(basic_total * float(EDLI_RATE), 2)
        pf_total = pf_emp + eps + epf_empl + edli

        # ESI figures
        esi_wages = float(
            payslips.filter(esi_employee__gt=0).aggregate(
                s=Sum("basic") + Sum("hra") + Sum("da") + Sum("other_allowances")
            )["s"] or 0
        )
        esi_emp_share  = round(esi_wages * float(ESI_EMP_RATE), 2)
        esi_empl_share = round(esi_wages * float(ESI_EMPLR_RATE), 2)
        esi_total = esi_emp_share + esi_empl_share

        # Challan status
        challans = {c.challan_type: c for c in run.challans.all()}
        pf_ch  = challans.get("epf")
        esi_ch = challans.get("esi")

        return Response({
            "payroll_run_id": run.id,
            "month_label":    f"{calendar.month_name[run.month]} {run.year}",
            "employees":      payslips.count(),
            "esi_eligible":   esi_eligible,

            "pf_employee_share":  pf_emp,
            "pf_employer_share":  epf_empl + eps + edli,
            "pf_eps":             eps,
            "pf_edli":            edli,
            "pf_total":           pf_total,
            "pf_generated":       pf_ch is not None,
            "pf_filed":           pf_ch.is_filed if pf_ch else False,

            "esi_wages":          esi_wages,
            "esi_employee_share": esi_emp_share,
            "esi_employer_share": esi_empl_share,
            "esi_total":          esi_total,
            "esi_generated":      esi_ch is not None,
            "esi_filed":          esi_ch.is_filed if esi_ch else False,
        })


# ── Generate challan ──────────────────────────────────────────
class GenerateChallanView(APIView):
    permission_classes = [IsAdminHR]

    def post(self, request):
        run_id = request.data.get("payroll_run")
        if not run_id:
            return Response({"detail": "payroll_run is required."}, status=400)
        try:
            run = PayrollRun.objects.get(pk=run_id)
        except PayrollRun.DoesNotExist:
            return Response({"detail": "PayrollRun not found."}, status=404)

        result = generate_compliance(run)
        return Response(result)


# ── Mark challan as filed ─────────────────────────────────────
class MarkFiledView(APIView):
    permission_classes = [IsAdminHR]

    def post(self, request, pk):
        try:
            challan = ChallanRun.objects.get(pk=pk)
        except ChallanRun.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)
        challan.is_filed = True
        challan.filed_on = date.today()
        challan.save()
        return Response(ChallanRunSerializer(challan).data)


# ── ECR (PF Challan) Excel ────────────────────────────────────
class EPFChallanExportView(APIView):
    permission_classes = [IsAdminHR]

    def get(self, request):
        run_id = request.query_params.get("payroll_run")
        if not run_id:
            return Response({"detail": "payroll_run is required."}, status=400)
        try:
            run = PayrollRun.objects.get(pk=run_id)
        except PayrollRun.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        payslips = run.payslips.select_related(
            "employee__site__district__state"
        ).order_by("employee__emp_code")

        wb = Workbook()
        ws = wb.active
        ws.title = "EPF ECR"

        mname = calendar.month_name[run.month]

        # Title
        ws.merge_cells("A1:L1")
        t = ws["A1"]
        t.value = f"EPF (PF) CHALLAN – ECR FORMAT · {mname.upper()} {run.year}"
        t.font  = Font(name="Calibri", bold=True, size=13, color="1E3563")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:L2")
        ws["A2"].value = f"Wage Month: {mname} {run.year}   |   Generated: {date.today():%d %b %Y}"
        ws["A2"].font  = Font(name="Calibri", size=9, color="6B7793")
        ws["A2"].alignment = Alignment(horizontal="center")

        HEADERS = [
            ("Sr",             5),  ("UAN",          16), ("Emp Code",     12),
            ("Name",          24),  ("Designation",  18), ("Gross Wages",  14),
            ("EPF Wages\n(Basic)",14),("EPS Wages",  14), ("Emp EPF\n(12%)",14),
            ("Empl EPS\n(8.33%)",14),("Empl EPF\n(3.67%)",14),("NCP Days",10),
        ]
        for ci, (h, w) in enumerate(HEADERS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            ws.cell(row=3, column=ci).value = h

        tot_gross = tot_epf_wages = tot_emp_epf = tot_empl_eps = tot_empl_epf = 0
        for ri, slip in enumerate(payslips, 4):
            emp  = slip.employee
            gross = float(slip.basic + slip.hra + slip.da + slip.other_allowances)
            basic = float(slip.basic)
            emp_epf   = round(basic * float(PF_EMP_RATE), 2)
            empl_eps  = round(min(basic * float(EPS_RATE), float(EPS_CAP)), 2)
            empl_epf  = round(basic * float(EPF_EMP_RATE), 2)
            ncp       = slip.working_days - slip.present_days

            ws.cell(row=ri, column=1).value  = ri - 3
            ws.cell(row=ri, column=2).value  = emp.uan or "Not Registered"
            ws.cell(row=ri, column=3).value  = emp.emp_code
            ws.cell(row=ri, column=4).value  = emp.full_name
            ws.cell(row=ri, column=5).value  = emp.designation
            ws.cell(row=ri, column=6).value  = round(gross, 2)
            ws.cell(row=ri, column=7).value  = round(basic, 2)
            ws.cell(row=ri, column=8).value  = round(basic, 2)   # EPS wages = basic
            ws.cell(row=ri, column=9).value  = emp_epf
            ws.cell(row=ri, column=10).value = empl_eps
            ws.cell(row=ri, column=11).value = empl_epf
            ws.cell(row=ri, column=12).value = ncp

            tot_gross     += gross
            tot_epf_wages += basic
            tot_emp_epf   += emp_epf
            tot_empl_eps  += empl_eps
            tot_empl_epf  += empl_epf

        # Totals row
        tr = payslips.count() + 4
        bf = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=tr, column=4,  value="TOTAL").font = bf
        for ci, val in zip([6,7,8,9,10,11],
                           [round(tot_gross,2), round(tot_epf_wages,2), round(tot_epf_wages,2),
                            round(tot_emp_epf,2), round(tot_empl_eps,2), round(tot_empl_epf,2)]):
            c = ws.cell(row=tr, column=ci, value=val)
            c.font = bf

        _tbl_style(ws, tr, 12, "1E3563")

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"epf_ecr_{run.year}_{run.month:02d}.xlsx"
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


# ── ESI Challan Excel ─────────────────────────────────────────
class ESIChallanExportView(APIView):
    permission_classes = [IsAdminHR]

    def get(self, request):
        run_id = request.query_params.get("payroll_run")
        if not run_id:
            return Response({"detail": "payroll_run is required."}, status=400)
        try:
            run = PayrollRun.objects.get(pk=run_id)
        except PayrollRun.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        payslips = run.payslips.select_related(
            "employee__site__district__state"
        ).filter(esi_employee__gt=0).order_by("employee__emp_code")

        wb = Workbook()
        ws = wb.active
        ws.title = "ESI Challan"

        mname = calendar.month_name[run.month]

        ws.merge_cells("A1:J1")
        t = ws["A1"]
        t.value = f"ESI CHALLAN · {mname.upper()} {run.year}"
        t.font  = Font(name="Calibri", bold=True, size=13, color="6A0DAD")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:J2")
        ws["A2"].value = f"Contribution Period: {mname} {run.year}   |   Generated: {date.today():%d %b %Y}"
        ws["A2"].font  = Font(name="Calibri", size=9, color="6B7793")
        ws["A2"].alignment = Alignment(horizontal="center")

        HEADERS = [
            ("Sr",             5),  ("ESIC No.",      16), ("Emp Code",    12),
            ("Employee Name", 24),  ("Designation",   18), ("IP Wages\n(Gross)",14),
            ("Emp ESI\n(0.75%)",14),("Empl ESI\n(3.25%)",14),("Total ESI",14),
            ("NCP Days",      10),
        ]
        for ci, (h, w) in enumerate(HEADERS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            ws.cell(row=3, column=ci).value = h

        tot_wages = tot_emp = tot_empl = 0
        for ri, slip in enumerate(payslips, 4):
            emp   = slip.employee
            gross = float(slip.basic + slip.hra + slip.da + slip.other_allowances)
            e_esi = round(gross * float(ESI_EMP_RATE), 2)
            r_esi = round(gross * float(ESI_EMPLR_RATE), 2)
            ncp   = slip.working_days - slip.present_days

            ws.cell(row=ri, column=1).value  = ri - 3
            ws.cell(row=ri, column=2).value  = emp.esic_no or "Not Registered"
            ws.cell(row=ri, column=3).value  = emp.emp_code
            ws.cell(row=ri, column=4).value  = emp.full_name
            ws.cell(row=ri, column=5).value  = emp.designation
            ws.cell(row=ri, column=6).value  = round(gross, 2)
            ws.cell(row=ri, column=7).value  = e_esi
            ws.cell(row=ri, column=8).value  = r_esi
            ws.cell(row=ri, column=9).value  = round(e_esi + r_esi, 2)
            ws.cell(row=ri, column=10).value = ncp

            tot_wages += gross; tot_emp += e_esi; tot_empl += r_esi

        tr = payslips.count() + 4
        bf = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=tr, column=4, value="TOTAL").font = bf
        for ci, val in zip([6,7,8,9],
                           [round(tot_wages,2), round(tot_emp,2),
                            round(tot_empl,2), round(tot_emp+tot_empl,2)]):
            ws.cell(row=tr, column=ci, value=val).font = bf

        _tbl_style(ws, tr, 10, "6A0DAD")

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"esi_challan_{run.year}_{run.month:02d}.xlsx"
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


# ── ViewSets (kept for browsable API) ────────────────────────
class PFContributionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PFContribution.objects.all()
    serializer_class = PFContributionSerializer
    permission_classes = [IsAdminHR]
    filterset_fields = ["payslip__payroll_run"]


class ESIContributionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ESIContribution.objects.all()
    serializer_class = ESIContributionSerializer
    permission_classes = [IsAdminHR]
    filterset_fields = ["payslip__payroll_run"]


class ChallanRunViewSet(viewsets.ModelViewSet):
    queryset = ChallanRun.objects.select_related("payroll_run").all()
    serializer_class = ChallanRunSerializer
    permission_classes = [IsAdminHR]
    filterset_fields = ["payroll_run", "challan_type", "is_filed"]
