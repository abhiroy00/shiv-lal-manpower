import io
import zipfile
import calendar
from decimal import Decimal
from datetime import date
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import SalaryStructure, PayrollRun, Payslip
from .serializers import SalaryStructureSerializer, PayrollRunSerializer, PayslipSerializer
from .services import run_payroll, PF_THRESHOLD
from .pdf_service import generate_payslip_pdf
from apps.common.permissions import IsAdminHR


class SalaryStructureViewSet(viewsets.ModelViewSet):
    queryset = SalaryStructure.objects.select_related("employee").all()
    serializer_class = SalaryStructureSerializer
    permission_classes = [IsAdminHR]
    filterset_fields  = ["employee"]

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        """Download a blank Excel template for bulk salary structure setup."""
        from apps.employees.models import Employee
        wb = Workbook()
        ws = wb.active
        ws.title = "Salary Structures"

        navy  = PatternFill("solid", fgColor="1E3563")
        gold  = PatternFill("solid", fgColor="D4AF37")
        nfont = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        gfont = Font(name="Calibri", bold=True, color="1E3563", size=10)
        ctr   = Alignment(horizontal="center", vertical="center")
        thin  = Side(style="thin", color="D0D7E5")
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title row
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = "SALARY STRUCTURE CONFIGURATION SHEET — M/S SHIV LAL MANPOWER"
        c.font  = Font(name="Calibri", bold=True, size=13, color="1E3563")
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 26

        # Rule note
        ws.merge_cells("A2:H2")
        ws["A2"].value = (
            "RULE: Basic ≥ ₹30,000 → TDS (10% of gross) deducted; PF & ESIC NOT deducted.  "
            "Basic < ₹30,000 → PF (12% of Basic) + ESIC (0.75% of Basic) deducted; TDS NOT deducted."
        )
        ws["A2"].font      = Font(name="Calibri", size=8.5, color="7B1FA2", italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

        COLS = [
            ("Emp Code",          14),
            ("Employee Name",     26),
            ("Basic (₹)",        14),
            ("HRA (₹)",          12),
            ("DA (₹)",           12),
            ("Other Allowances (₹)", 20),
            ("Regime (auto)",    16),
            ("Notes",            30),
        ]
        ws.row_dimensions[3].height = 20
        for ci, (label, w) in enumerate(COLS, 1):
            c = ws.cell(row=3, column=ci, value=label)
            c.font = nfont; c.fill = navy; c.alignment = ctr; c.border = bdr
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Pre-fill with existing employees
        employees = Employee.objects.filter(status="active").select_related("salary_structure").order_by("emp_code")
        cfont = Font(name="Calibri", size=10)
        for ri, emp in enumerate(employees, 4):
            try:
                ss = emp.salary_structure
                basic = float(ss.basic); hra = float(ss.hra); da = float(ss.da); other = float(ss.other_allowances)
            except SalaryStructure.DoesNotExist:
                basic = hra = da = other = 0.0
            regime = "TDS" if basic >= float(PF_THRESHOLD) else "PF + ESIC"
            row = [emp.emp_code, emp.full_name, basic, hra, da, other, regime, ""]
            alt  = PatternFill("solid", fgColor="F4F6FA") if ri % 2 == 0 else None
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = cfont; c.border = bdr
                if alt: c.fill = alt
                if ci in (3, 4, 5, 6) and val:
                    c.number_format = '₹#,##0.00'

        ws.freeze_panes = "A4"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="salary_structure_template.xlsx"'
        return resp

    @action(detail=False, methods=["post"], url_path="upload")
    def upload_structures(self, request):
        """Bulk create/update salary structures from uploaded Excel."""
        from apps.employees.models import Employee
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file uploaded."}, status=400)
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
        except Exception:
            return Response({"detail": "Invalid Excel file."}, status=400)

        created = updated = errors = 0
        error_list = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            emp_code = str(row[0] or "").strip()
            if not emp_code:
                continue
            try:
                basic = Decimal(str(row[2] or 0))
                hra   = Decimal(str(row[3] or 0))
                da    = Decimal(str(row[4] or 0))
                other = Decimal(str(row[5] or 0))
            except Exception:
                error_list.append(f"{emp_code}: invalid numeric values")
                errors += 1
                continue
            try:
                emp = Employee.objects.get(emp_code=emp_code)
            except Employee.DoesNotExist:
                error_list.append(f"{emp_code}: employee not found")
                errors += 1
                continue
            _, was_created = SalaryStructure.objects.update_or_create(
                employee=emp,
                defaults={"basic": basic, "hra": hra, "da": da, "other_allowances": other},
            )
            if was_created:
                created += 1
            else:
                updated += 1

        return Response({"created": created, "updated": updated, "errors": errors, "error_list": error_list})


class PayrollRunViewSet(viewsets.ModelViewSet):
    queryset = PayrollRun.objects.all()
    serializer_class = PayrollRunSerializer
    permission_classes = [IsAdminHR]

    @action(detail=False, methods=["post"])
    def run(self, request):
        month = request.data.get("month")
        year  = request.data.get("year")
        if not month or not year:
            return Response({"detail": "month and year are required."}, status=400)
        payroll_run, stats = run_payroll(int(month), int(year), request.user)
        data = PayrollRunSerializer(payroll_run).data
        data["stats"] = stats
        return Response(data)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        run = self.get_object()
        if run.run_status != PayrollRun.RunStatus.DRAFT:
            return Response({"detail": "Only draft payrolls can be approved."}, status=400)
        run.run_status = PayrollRun.RunStatus.APPROVED
        run.save()
        return Response(PayrollRunSerializer(run).data)

    @action(detail=True, methods=["post"], url_path="mark-paid")
    def mark_paid(self, request, pk=None):
        run = self.get_object()
        if run.run_status != PayrollRun.RunStatus.APPROVED:
            return Response({"detail": "Only approved payrolls can be marked paid."}, status=400)
        run.run_status = PayrollRun.RunStatus.PAID
        run.save()
        return Response(PayrollRunSerializer(run).data)

    @action(detail=True, methods=["get"], url_path="bank-advice")
    def bank_advice(self, request, pk=None):
        run = self.get_object()
        payslips = run.payslips.select_related(
            "employee__site__district__state"
        ).order_by("employee__emp_code")

        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Advice"

        thin   = Side(style="thin", color="D0D7E5")
        bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr    = Alignment(horizontal="center", vertical="center")
        navy   = PatternFill("solid", fgColor="1E3563")
        nfont  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cfont  = Font(name="Calibri", size=10)
        bfont  = Font(name="Calibri", bold=True, size=10)
        mname  = calendar.month_name[run.month]

        # Title
        ws.merge_cells("A1:H1")
        t = ws["A1"]
        t.value     = f"BANK ADVICE – {mname.upper()} {run.year}"
        t.font      = Font(name="Calibri", bold=True, size=14, color="1E3563")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:H2")
        ws["A2"].value = f"Status: {run.run_status.upper()}   |   Generated: {date.today().strftime('%d %b %Y')}"
        ws["A2"].font  = Font(name="Calibri", size=9, color="6B7793")
        ws["A2"].alignment = Alignment(horizontal="center")

        HEADERS = [
            ("Sr",            6),
            ("Emp Code",     12),
            ("Employee Name", 26),
            ("Designation",  18),
            ("Site",         20),
            ("Bank Account", 20),
            ("IFSC Code",    14),
            ("Net Pay (₹)",  14),
        ]

        ws.row_dimensions[3].height = 20
        for ci, (label, width) in enumerate(HEADERS, start=1):
            c = ws.cell(row=3, column=ci, value=label)
            c.font = nfont; c.fill = navy; c.alignment = ctr; c.border = bdr
            ws.column_dimensions[get_column_letter(ci)].width = width

        total_net = 0
        for ri, slip in enumerate(payslips, start=4):
            alt  = (ri % 2 == 0)
            fill = PatternFill("solid", fgColor="F4F6FA") if alt else None
            emp  = slip.employee

            row_vals = [
                ri - 3,
                emp.emp_code,
                emp.full_name,
                emp.designation,
                emp.site.name if emp.site else "",
                emp.bank_account or "—",
                emp.ifsc        or "—",
                float(slip.net_pay),
            ]
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = cfont; c.border = bdr
                if fill: c.fill = fill
                if ci == 1 or ci == 8: c.alignment = ctr
                if ci == 8:
                    c.number_format = '₹#,##0.00'
                    c.font = bfont
            total_net += float(slip.net_pay)

        # Totals row
        tot_row = len(payslips) + 4
        ws.cell(row=tot_row, column=7, value="TOTAL").font = bfont
        tc = ws.cell(row=tot_row, column=8, value=total_net)
        tc.font = Font(name="Calibri", bold=True, size=11, color="15966A")
        tc.number_format = '₹#,##0.00'
        tc.border = bdr; tc.alignment = ctr

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:H3"

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"bank_advice_{run.year}_{run.month:02d}.xlsx"
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

    @action(detail=True, methods=["get"], url_path="salary-sheet")
    def salary_sheet(self, request, pk=None):
        """Comprehensive salary sheet Excel for a payroll run."""
        run = self.get_object()
        payslips = run.payslips.select_related(
            "employee__site__district__state"
        ).order_by("employee__emp_code")

        wb    = Workbook()
        ws    = wb.active
        ws.title = "Salary Sheet"
        mname = calendar.month_name[run.month]

        thin  = Side(style="thin", color="D0D7E5")
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr   = Alignment(horizontal="center", vertical="center")
        navy  = PatternFill("solid", fgColor="1E3563")
        tds_f = PatternFill("solid", fgColor="F3E5F5")
        pf_f  = PatternFill("solid", fgColor="E3F2FD")
        nfont = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cfont = Font(name="Calibri", size=10)
        bfont = Font(name="Calibri", bold=True, size=10)

        # Title
        ws.merge_cells("A1:P1")
        t = ws["A1"]
        t.value     = f"SALARY SHEET — {mname.upper()} {run.year} — M/S SHIV LAL MANPOWER"
        t.font      = Font(name="Calibri", bold=True, size=13, color="1E3563")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:P2")
        ws["A2"].value = (
            f"Status: {run.run_status.upper()}  |  Generated: {date.today().strftime('%d %b %Y')}  |  "
            f"Rule: Basic > ₹30,000 → TDS (10%)  |  Basic ≤ ₹30,000 → PF(12%) + ESIC(0.75%) + Bonus(8.33%)"
        )
        ws["A2"].font      = Font(name="Calibri", size=8.5, color="6B7793", italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 16

        # 16 columns — other_allowances removed (dormant field, always ₹0 per policy)
        HEADERS = [
            ("Sr",              5),  ("Emp Code",    12), ("Employee Name",  26),
            ("Designation",    18),  ("Site",         18), ("Days P/W",       9),
            ("Basic (₹)",     12),  ("HRA (₹)",     10), ("DA (₹)",         10),
            ("Bonus (₹)",     11),  ("Gross CTC (₹)",13),
            ("PF Emp (₹)",    11),  ("ESIC Emp (₹)",11), ("TDS (₹)",        11),
            ("Other Ded (₹)", 11),  ("Net Pay (₹)", 13),
        ]
        ws.row_dimensions[3].height = 20
        for ci, (label, w) in enumerate(HEADERS, 1):
            c = ws.cell(row=3, column=ci, value=label)
            c.font = nfont; c.fill = navy; c.alignment = ctr; c.border = bdr
            ws.column_dimensions[get_column_letter(ci)].width = w

        total_gross = total_pf = total_esi = total_tds = total_net = 0.0
        for ri, slip in enumerate(payslips, 4):
            emp    = slip.employee
            pf_mode = float(slip.bonus) > 0   # bonus > 0 → PF/ESIC regime
            is_tds  = not pf_mode
            alt    = PatternFill("solid", fgColor="F4F6FA") if ri % 2 == 0 else None

            basic   = float(slip.basic)
            hra     = float(slip.hra)
            da      = float(slip.da)
            bonus   = float(slip.bonus)
            pf_emp  = float(slip.pf_employee)
            esi_emp = float(slip.esi_employee)
            tds     = float(slip.tds)
            other_d = float(slip.other_deductions)

            if pf_mode:
                pf_er  = round(basic * 0.12,   2)
                esi_er = round(basic * 0.0325,  2)
                gross  = basic + hra + da + pf_er + esi_er + bonus
                net    = round(gross - (pf_emp + pf_er + esi_emp + esi_er + other_d), 2)
            else:
                pf_er = esi_er = 0.0
                gross = basic + hra + da
                net   = round(gross - (tds + other_d), 2)

            row_vals = [
                ri - 3,
                emp.emp_code,
                emp.full_name,
                emp.designation,
                emp.site.name if emp.site else "",
                f"{slip.present_days}/{slip.working_days}",
                basic,
                hra,
                da,
                bonus,
                gross,
                pf_emp,
                esi_emp,
                tds,
                other_d,
                net,
            ]
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = bdr
                c.font   = bfont if ci in (11, 16) else cfont
                if alt:   c.fill = alt
                if ci == 1: c.alignment = ctr
                if ci >= 7:
                    c.number_format = '₹#,##0.00'
                if ci in (12, 13) and pf_mode:
                    c.fill = pf_f
                if ci == 14 and is_tds:
                    c.fill = tds_f

            total_gross += gross
            total_pf    += pf_emp
            total_esi   += esi_emp
            total_tds   += tds
            total_net   += net

        # Totals row
        tot = len(payslips) + 4
        ws.cell(tot, 2, "TOTAL").font = bfont
        for ci, val in [(11, total_gross),(12, total_pf),(13, total_esi),(14, total_tds),(16, total_net)]:
            c = ws.cell(tot, ci, val)
            c.font   = Font(name="Calibri", bold=True, size=10, color="15966A")
            c.number_format = '₹#,##0.00'
            c.border = bdr

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"salary_sheet_{run.year}_{run.month:02d}.xlsx"
        resp  = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

    @action(detail=True, methods=["get"], url_path="payslips-zip")
    def payslips_zip(self, request, pk=None):
        """Bulk download all payslips for a run as a ZIP of PDFs."""
        run = self.get_object()
        payslips = run.payslips.select_related(
            "employee__site__district__state"
        ).order_by("employee__emp_code")

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for slip in payslips:
                pdf_buf = generate_payslip_pdf(slip)
                fname   = f"{slip.employee.emp_code}_{run.year}_{run.month:02d}.pdf"
                zf.writestr(fname, pdf_buf.read())

        buf.seek(0)
        zip_name = f"payslips_{run.year}_{run.month:02d}.zip"
        resp = HttpResponse(buf.read(), content_type="application/zip")
        resp["Content-Disposition"] = f'attachment; filename="{zip_name}"'
        return resp


class PayslipViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Payslip.objects.select_related(
        "employee__site__district__state", "payroll_run"
    ).all()
    serializer_class = PayslipSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["payroll_run", "employee"]
    search_fields = ["employee__full_name", "employee__emp_code", "employee__designation"]
    ordering = ["employee__emp_code"]
    pagination_class = None  # always filtered by payroll_run, bounded per run

    @action(detail=True, methods=["get"], url_path="pdf")
    def pdf(self, request, pk=None):
        slip     = self.get_object()
        pdf_buf  = generate_payslip_pdf(slip)
        emp_code = slip.employee.emp_code
        fname    = f"payslip_{emp_code}_{slip.payroll_run.year}_{slip.payroll_run.month:02d}.pdf"
        resp = HttpResponse(pdf_buf.read(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        # Never allow browser/proxy to cache payslip PDFs — data changes after re-runs
        resp["Cache-Control"] = "no-cache, no-store, must-revalidate"
        resp["Pragma"]        = "no-cache"
        resp["Expires"]       = "0"
        return resp


class MyPayslipsView(APIView):
    """Employee: own payslips list + optional detail."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        emp = request.user.employee
        if emp is None:
            return Response({"detail": "No employee linked."}, status=400)
        # Prefetch site to avoid extra query on emp.site.name
        from apps.employees.models import Employee as EmpModel
        emp = EmpModel.objects.select_related("site").get(pk=emp.pk)
        slips = (
            Payslip.objects
            .filter(employee=emp)
            .select_related("payroll_run")
            .order_by("-payroll_run__year", "-payroll_run__month")
        )
        MONTH_NAMES = ["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        from .serializers import _compute_figures
        data = []
        for s in slips:
            m = s.payroll_run.month
            y = s.payroll_run.year
            gross_ctc, total_d, net, pf_er, esi_er = _compute_figures(s)
            data.append({
                "id":               s.id,
                "month":            m,
                "year":             y,
                "run_month":        m,
                "run_year":         y,
                "month_label":      f"{MONTH_NAMES[m]} {y}" if 1 <= m <= 12 else f"{m}/{y}",
                "run_status":       s.payroll_run.run_status,
                "emp_code":         emp.emp_code,
                "employee_name":    emp.full_name,
                "designation":      emp.designation,
                "site_name":        emp.site.name if emp.site else "",
                "bank_account":     emp.bank_account or "",
                "ifsc":             emp.ifsc or "",
                "present_days":     s.present_days,
                "working_days":     s.working_days,
                "basic":            str(s.basic),
                "hra":              str(s.hra),
                "da":               str(s.da),
                "bonus":            str(s.bonus),
                "pf_employee":      str(s.pf_employee),
                "esi_employee":     str(s.esi_employee),
                "pf_employer":      str(pf_er),
                "esi_employer":     str(esi_er),
                "tds":              str(s.tds),
                "other_deductions": str(s.other_deductions),
                "gross_pay":        str(gross_ctc),
                "net_pay":          str(net),
                "pdf_url":          request.build_absolute_uri(f"/api/payslips/{s.id}/pdf/"),
            })
        response = Response(data)
        response["Cache-Control"] = "no-cache, no-store, must-revalidate"
        return response
