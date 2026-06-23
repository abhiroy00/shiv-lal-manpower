"""
Report views — each endpoint generates and streams an Excel file.
All reports use openpyxl with consistent styling.
"""
import io
import calendar
from datetime import date
from django.http import HttpResponse
from django.db.models import Count, Sum, Q, Avg
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from apps.common.permissions import IsAdminHR
from apps.employees.models import Employee
from apps.attendance.models import Attendance
from apps.deployment.models import Site, District, State
from apps.payroll.models import PayrollRun, Payslip
from apps.recruitment.models import Candidate, Requisition

# ── Shared style helpers ──────────────────────────────────────
THIN = Side(style="thin", color="D0D7E5")
BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR  = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left",   vertical="center")


def _hdr_style(cell, hex_bg="1E3563"):
    cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", fgColor=hex_bg)
    cell.alignment = CTR
    cell.border    = BDR


def _title_row(ws, text, ncols, hex_color="1E3563"):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font      = Font(name="Calibri", bold=True, size=13, color=hex_color)
    c.alignment = CTR
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1).value = f"Generated: {date.today():%d %b %Y}"
    ws.cell(row=2, column=1).font  = Font(name="Calibri", size=9, color="6B7793")
    ws.cell(row=2, column=1).alignment = CTR


def _apply_table(ws, hdr_row, last_row, ncols, hdr_hex="1E3563"):
    for col in range(1, ncols + 1):
        _hdr_style(ws.cell(row=hdr_row, column=col), hdr_hex)

    alt = PatternFill("solid", fgColor="F4F6FA")
    cfont = Font(name="Calibri", size=10)
    for r in range(hdr_row + 1, last_row + 1):
        is_alt = (r % 2 == 0)
        for col in range(1, ncols + 1):
            c = ws.cell(row=r, column=col)
            c.font   = cfont
            c.border = BDR
            if is_alt:
                c.fill = alt

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ── 1. Payroll Statement ──────────────────────────────────────
class PayrollStatementView(APIView):
    permission_classes = [IsAdminHR]

    def get(self, request):
        month = int(request.query_params.get("month", date.today().month))
        year  = int(request.query_params.get("year",  date.today().year))
        run   = PayrollRun.objects.filter(month=month, year=year).first()

        wb = Workbook(); ws = wb.active; ws.title = "Payroll Statement"
        mname = calendar.month_name[month]
        _title_row(ws, f"PAYROLL STATEMENT — {mname.upper()} {year}", 14)

        COLS = [
            ("Sr",4),("Emp Code",11),("Name",22),("Designation",16),("Site",18),
            ("Basic",12),("HRA",11),("DA",11),("Other Allow.",11),
            ("Gross",12),("PF (Emp)",11),("ESI (Emp)",11),("Other Ded.",11),("Net Pay",12),
        ]
        for ci,(h,w) in enumerate(COLS,1):
            ws.cell(row=3,column=ci).value=h
            ws.column_dimensions[get_column_letter(ci)].width=w
        ws.row_dimensions[3].height=20

        payslips = (run.payslips.select_related("employee__site")
                    .order_by("employee__emp_code") if run else [])

        tot = dict(gross=0,pf=0,esi=0,ded=0,net=0)
        for ri,p in enumerate(payslips,4):
            e=p.employee
            gross=float(p.basic+p.hra+p.da+p.other_allowances)
            vals=[ri-3,e.emp_code,e.full_name,e.designation,
                  e.site.name if e.site else "",
                  float(p.basic),float(p.hra),float(p.da),float(p.other_allowances),
                  gross,float(p.pf_employee),float(p.esi_employee),
                  float(p.other_deductions),float(p.net_pay)]
            for ci,v in enumerate(vals,1):
                ws.cell(row=ri,column=ci).value=v
            tot["gross"]+=gross; tot["pf"]+=float(p.pf_employee)
            tot["esi"]+=float(p.esi_employee); tot["net"]+=float(p.net_pay)

        tr=len(list(payslips))+4
        bf=Font(name="Calibri",bold=True,size=10,color="15966A")
        ws.cell(row=tr,column=4,value="TOTAL").font=Font(name="Calibri",bold=True,size=10)
        for ci,k in [(10,"gross"),(11,"pf"),(12,"esi"),(14,"net")]:
            c=ws.cell(row=tr,column=ci,value=round(tot[k],2))
            c.font=bf; c.border=BDR

        _apply_table(ws,3,tr,14)
        ws.auto_filter.ref="A3:N3"
        return _xlsx_response(wb,f"payroll_statement_{year}_{month:02d}.xlsx")


# ── 2. Monthly Attendance Summary ─────────────────────────────
class AttendanceSummaryView(APIView):
    permission_classes = [IsAdminHR]

    def get(self, request):
        month   = int(request.query_params.get("month", date.today().month))
        year    = int(request.query_params.get("year",  date.today().year))
        site_id = request.query_params.get("site")

        days_in_month = calendar.monthrange(year,month)[1]
        sundays = sum(1 for d in range(1,days_in_month+1)
                      if calendar.weekday(year,month,d)==6)
        working_days = days_in_month - sundays

        emps = Employee.objects.filter(status="active").select_related("site")
        if site_id:
            emps = emps.filter(site_id=site_id)
        emps = emps.order_by("emp_code")

        start = date(year,month,1)
        end   = date(year,month,days_in_month)
        att_qs = Attendance.objects.filter(
            date__range=(start,end), employee__in=emps
        ).values("employee_id","status").annotate(n=Count("id"))

        att_map={}
        for row in att_qs:
            eid=row["employee_id"]
            if eid not in att_map:
                att_map[eid]={"present":0,"late":0,"absent":0}
            if row["status"]=="present": att_map[eid]["present"]+=row["n"]
            elif row["status"]=="late":  att_map[eid]["late"]+=row["n"]
            elif row["status"]=="absent":att_map[eid]["absent"]+=row["n"]

        wb=Workbook(); ws=wb.active; ws.title="Attendance Summary"
        mname=calendar.month_name[month]
        _title_row(ws,f"MONTHLY ATTENDANCE SUMMARY — {mname.upper()} {year}",11,"15966A")
        ws.merge_cells("A2:K2")
        ws.cell(row=2,column=1).value=(
            f"Working Days: {working_days} | Generated: {date.today():%d %b %Y}"
        )
        ws.cell(row=2,column=1).font=Font(name="Calibri",size=9,color="6B7793")
        ws.cell(row=2,column=1).alignment=CTR

        COLS=[("Sr",4),("Emp Code",11),("Name",22),("Designation",16),("Site",18),
              ("Total Days",9),("Present",9),("Late",9),("Absent",9),
              ("Sunday",8),("Att %",8)]
        for ci,(h,w) in enumerate(COLS,1):
            ws.cell(row=3,column=ci).value=h
            ws.column_dimensions[get_column_letter(ci)].width=w
        ws.row_dimensions[3].height=20

        GREEN=PatternFill("solid",fgColor="E1F4EC")
        RED=PatternFill("solid",fgColor="FDECEA")
        YEL=PatternFill("solid",fgColor="FBF1DC")

        for ri,emp in enumerate(emps,4):
            a=att_map.get(emp.id,{"present":0,"late":0,"absent":0})
            pct=round((a["present"]+a["late"])/working_days*100,1) if working_days else 0
            vals=[ri-3,emp.emp_code,emp.full_name,emp.designation,
                  emp.site.name if emp.site else "",
                  working_days,a["present"],a["late"],a["absent"],sundays,pct]
            for ci,v in enumerate(vals,1):
                ws.cell(row=ri,column=ci).value=v
            # Colour att% cell
            pct_cell=ws.cell(row=ri,column=11)
            if pct>=90:   pct_cell.fill=GREEN
            elif pct>=75: pct_cell.fill=YEL
            else:         pct_cell.fill=RED

        tr=emps.count()+4
        bf=Font(name="Calibri",bold=True,size=10)
        ws.cell(row=tr,column=3,value="TOTAL").font=bf
        _apply_table(ws,3,tr,11,"15966A")
        ws.auto_filter.ref="A3:K3"
        return _xlsx_response(wb,f"attendance_summary_{year}_{month:02d}.xlsx")


# ── 3. Deployment Strength ────────────────────────────────────
class DeploymentStrengthView(APIView):
    permission_classes = [IsAdminHR]

    def get(self, request):
        sites=Site.objects.filter(is_active=True).select_related(
            "district__state"
        ).annotate(deployed=Count("employees",filter=Q(employees__status="active"))
        ).order_by("district__state__name","district__name","name")

        wb=Workbook(); ws=wb.active; ws.title="Deployment Strength"
        _title_row(ws,"SITE-WISE DEPLOYMENT STRENGTH REPORT",9,"6A0DAD")

        COLS=[("Sr",4),("State",14),("District",14),("Site",24),
              ("Sanctioned",10),("Deployed",10),("Vacancy",10),("Fill %",8),("Status",12)]
        for ci,(h,w) in enumerate(COLS,1):
            ws.cell(row=3,column=ci).value=h
            ws.column_dimensions[get_column_letter(ci)].width=w
        ws.row_dimensions[3].height=20

        GREEN=PatternFill("solid",fgColor="E1F4EC")
        YEL  =PatternFill("solid",fgColor="FBF1DC")
        RED  =PatternFill("solid",fgColor="FDECEA")
        GFONT=Font(name="Calibri",size=10,color="15966A",bold=True)
        RFONT=Font(name="Calibri",size=10,color="D2453F",bold=True)
        YFONT=Font(name="Calibri",size=10,color="C98A12",bold=True)

        tot_sanc=tot_dep=0
        for ri,s in enumerate(sites,4):
            vac=s.sanctioned_strength-s.deployed
            pct=round(s.deployed/s.sanctioned_strength*100,1) if s.sanctioned_strength else 0
            status="Full" if pct>=90 else ("Partial" if pct>=60 else "Critical")
            vals=[ri-3,s.district.state.name,s.district.name,s.name,
                  s.sanctioned_strength,s.deployed,vac,pct,status]
            for ci,v in enumerate(vals,1):
                ws.cell(row=ri,column=ci).value=v
            # colour status cell
            sc=ws.cell(row=ri,column=9)
            if pct>=90:    sc.fill=GREEN; sc.font=GFONT
            elif pct>=60:  sc.fill=YEL;   sc.font=YFONT
            else:          sc.fill=RED;    sc.font=RFONT
            tot_sanc+=s.sanctioned_strength; tot_dep+=s.deployed

        tr=sites.count()+4
        bf=Font(name="Calibri",bold=True,size=10)
        ws.cell(row=tr,column=4,value="TOTAL").font=bf
        ws.cell(row=tr,column=5,value=tot_sanc).font=bf
        ws.cell(row=tr,column=6,value=tot_dep).font=bf
        ws.cell(row=tr,column=7,value=tot_sanc-tot_dep).font=bf
        fill_overall=round(tot_dep/tot_sanc*100,1) if tot_sanc else 0
        ws.cell(row=tr,column=8,value=fill_overall).font=bf

        _apply_table(ws,3,tr,9,"6A0DAD")
        ws.auto_filter.ref="A3:I3"
        return _xlsx_response(wb,"deployment_strength.xlsx")


# ── 4. PF/ESI Deduction Register ─────────────────────────────
class DeductionRegisterView(APIView):
    permission_classes = [IsAdminHR]

    def get(self, request):
        month=int(request.query_params.get("month",date.today().month))
        year =int(request.query_params.get("year", date.today().year))
        run  =PayrollRun.objects.filter(month=month,year=year).first()

        wb=Workbook(); ws=wb.active; ws.title="Deduction Register"
        mname=calendar.month_name[month]
        _title_row(ws,f"PF / ESI DEDUCTION REGISTER — {mname.upper()} {year}",13,"D2453F")

        COLS=[("Sr",4),("Emp Code",11),("Name",22),("UAN",16),
              ("Basic",12),("PF Emp\n(12%)",10),("EPS\n(8.33%)",10),("EPF Empl\n(3.67%)",10),
              ("ESI Emp\n(0.75%)",10),("ESI Empl\n(3.25%)",10),
              ("Total PF",12),("Total ESI",12),("Grand Total",12)]
        for ci,(h,w) in enumerate(COLS,1):
            ws.cell(row=3,column=ci).value=h
            ws.column_dimensions[get_column_letter(ci)].width=w
        ws.row_dimensions[3].height=28

        payslips=(run.payslips.select_related("employee")
                  .order_by("employee__emp_code") if run else [])

        from decimal import Decimal
        EPS_RATE=Decimal("0.0833"); EPF_EMPL=Decimal("0.0367")
        EPS_CAP=Decimal("1250"); ESI_EMPL=Decimal("0.0325")

        tot={k:0 for k in ["pf_emp","eps","epf_empl","esi_emp","esi_empl","tot_pf","tot_esi"]}
        for ri,p in enumerate(payslips,4):
            e=p.employee
            basic=p.basic
            eps=float(min(basic*EPS_RATE,EPS_CAP))
            epf_empl=float(basic*EPF_EMPL)
            gross=float(p.basic+p.hra+p.da+p.other_allowances)
            esi_empl=round(gross*float(ESI_EMPL),2) if p.esi_employee>0 else 0
            tot_pf=float(p.pf_employee)+eps+epf_empl
            tot_esi=float(p.esi_employee)+esi_empl
            vals=[ri-3,e.emp_code,e.full_name,e.uan or "—",
                  float(basic),float(p.pf_employee),round(eps,2),round(epf_empl,2),
                  float(p.esi_employee),round(esi_empl,2),
                  round(tot_pf,2),round(tot_esi,2),round(tot_pf+tot_esi,2)]
            for ci,v in enumerate(vals,1):
                ws.cell(row=ri,column=ci).value=v
            for k,v in [("pf_emp",float(p.pf_employee)),("eps",eps),("epf_empl",epf_empl),
                         ("esi_emp",float(p.esi_employee)),("esi_empl",esi_empl),
                         ("tot_pf",tot_pf),("tot_esi",tot_esi)]:
                tot[k]+=v

        tr=len(list(payslips))+4
        bf=Font(name="Calibri",bold=True,size=10,color="D2453F")
        ws.cell(row=tr,column=4,value="TOTAL").font=Font(name="Calibri",bold=True,size=10)
        for ci,k in [(6,"pf_emp"),(7,"eps"),(8,"epf_empl"),
                     (9,"esi_emp"),(10,"esi_empl"),(11,"tot_pf"),(12,"tot_esi")]:
            c=ws.cell(row=tr,column=ci,value=round(tot[k],2))
            c.font=bf; c.border=BDR
        _apply_table(ws,3,tr,13,"D2453F")
        ws.auto_filter.ref="A3:M3"
        return _xlsx_response(wb,f"deduction_register_{year}_{month:02d}.xlsx")


# ── 5. Recruitment Status ─────────────────────────────────────
class RecruitmentStatusView(APIView):
    permission_classes = [IsAdminHR]

    def get(self, request):
        wb=Workbook()

        # Sheet 1 — Pipeline summary
        ws1=wb.active; ws1.title="Pipeline Summary"
        _title_row(ws1,"RECRUITMENT PIPELINE SUMMARY",4,"E8821E")
        ws1.cell(row=3,column=1).value="Stage"
        ws1.cell(row=3,column=2).value="Count"
        ws1.cell(row=3,column=3).value="Share %"
        ws1.cell(row=3,column=4).value="Status"
        for w,col in zip([18,10,10,14],[1,2,3,4]):
            ws1.column_dimensions[get_column_letter(col)].width=w
        ws1.row_dimensions[3].height=20

        from apps.recruitment.models import Candidate as C
        from django.db.models import Count as Cnt
        counts=dict(C.objects.values_list("stage").annotate(n=Cnt("id")))
        total=sum(counts.values()) or 1
        STAGES=[("applied","Applied"),("screened","Screened"),
                ("interview","Interview"),("selected","Selected"),("rejected","Rejected")]
        FILLS={"applied":"F0F2F8","screened":"EBF3FC","interview":"FBF1DC",
               "selected":"E1F4EC","rejected":"FDECEA"}
        for ri,(s,label) in enumerate(STAGES,4):
            n=counts.get(s,0)
            pct=round(n/total*100,1)
            ws1.cell(row=ri,column=1).value=label
            ws1.cell(row=ri,column=2).value=n
            ws1.cell(row=ri,column=3).value=pct
            ws1.cell(row=ri,column=4).value="Active" if s not in ("selected","rejected") else s.capitalize()
            for col in range(1,5):
                ws1.cell(row=ri,column=col).fill=PatternFill("solid",fgColor=FILLS[s])
        _apply_table(ws1,3,9,4,"E8821E")

        # Sheet 2 — All candidates
        ws2=wb.create_sheet("Candidates")
        _title_row(ws2,"ALL CANDIDATES",7,"E8821E")
        COLS2=[("Sr",4),("Name",22),("Phone",13),("Designation",16),
               ("Experience",11),("Stage",12),("Requisition",24)]
        for ci,(h,w) in enumerate(COLS2,1):
            ws2.cell(row=3,column=ci).value=h
            ws2.column_dimensions[get_column_letter(ci)].width=w
        ws2.row_dimensions[3].height=20
        cands=C.objects.select_related("requisition__site").order_by("stage","full_name")
        for ri,c in enumerate(cands,4):
            lbl=f"{c.requisition.designation} – {c.requisition.site.name}" if c.requisition else "—"
            ws2.cell(row=ri,column=1).value=ri-3
            ws2.cell(row=ri,column=2).value=c.full_name
            ws2.cell(row=ri,column=3).value=c.phone
            ws2.cell(row=ri,column=4).value=c.designation
            ws2.cell(row=ri,column=5).value=c.experience_years
            ws2.cell(row=ri,column=6).value=c.get_stage_display()
            ws2.cell(row=ri,column=7).value=lbl
        _apply_table(ws2,3,cands.count()+4,7,"E8821E")

        # Sheet 3 — Requisitions
        ws3=wb.create_sheet("Requisitions")
        _title_row(ws3,"OPEN REQUISITIONS",5,"E8821E")
        COLS3=[("Sr",4),("Site",22),("District",16),("Designation",18),("Vacancies",10)]
        for ci,(h,w) in enumerate(COLS3,1):
            ws3.cell(row=3,column=ci).value=h
            ws3.column_dimensions[get_column_letter(ci)].width=w
        reqs=Requisition.objects.filter(is_open=True).select_related("site__district")
        for ri,r in enumerate(reqs,4):
            ws3.cell(row=ri,column=1).value=ri-3
            ws3.cell(row=ri,column=2).value=r.site.name
            ws3.cell(row=ri,column=3).value=r.site.district.name
            ws3.cell(row=ri,column=4).value=r.designation
            ws3.cell(row=ri,column=5).value=r.count_required
        _apply_table(ws3,3,reqs.count()+4,5,"E8821E")

        return _xlsx_response(wb,"recruitment_status.xlsx")


# ── 6. MIS Pack (multi-sheet) ─────────────────────────────────
class MISPackView(APIView):
    permission_classes = [IsAdminHR]

    def get(self, request):
        month=int(request.query_params.get("month",date.today().month))
        year =int(request.query_params.get("year", date.today().year))
        mname=calendar.month_name[month]

        wb=Workbook()

        # Sheet 1 — Headcount by designation
        ws1=wb.active; ws1.title="Headcount"
        _title_row(ws1,f"HEADCOUNT SUMMARY — {mname.upper()} {year}",4)
        ws1.cell(row=3,column=1).value="Designation"
        ws1.cell(row=3,column=2).value="Active"
        ws1.cell(row=3,column=3).value="On Leave"
        ws1.cell(row=3,column=4).value="Inactive"
        for w,col in zip([22,10,10,10],[1,2,3,4]):
            ws1.column_dimensions[get_column_letter(col)].width=w
        from django.db.models import Count as Cnt
        desg=(Employee.objects.values("designation","status")
              .annotate(n=Cnt("id")).order_by("designation"))
        desg_map={}
        for row in desg:
            d=row["designation"]
            if d not in desg_map: desg_map[d]={"active":0,"on_leave":0,"inactive":0}
            s=row["status"]; v=row["n"]
            if s=="active": desg_map[d]["active"]+=v
            elif s=="on_leave": desg_map[d]["on_leave"]+=v
            else: desg_map[d]["inactive"]+=v
        for ri,(d,vals) in enumerate(sorted(desg_map.items()),4):
            ws1.cell(row=ri,column=1).value=d
            ws1.cell(row=ri,column=2).value=vals["active"]
            ws1.cell(row=ri,column=3).value=vals["on_leave"]
            ws1.cell(row=ri,column=4).value=vals["inactive"]
        _apply_table(ws1,3,len(desg_map)+4,4)

        # Sheet 2 — Site deployment
        ws2=wb.create_sheet("Deployment")
        _title_row(ws2,"SITE DEPLOYMENT STATUS",6,"6A0DAD")
        for ci,(h,w) in enumerate(
            [("Sr",4),("State",14),("District",14),("Site",22),("Sanctioned",10),("Deployed",10)],1):
            ws2.cell(row=3,column=ci).value=h
            ws2.column_dimensions[get_column_letter(ci)].width=w
        sites=Site.objects.filter(is_active=True).select_related("district__state").annotate(
            dep=Cnt("employees",filter=Q(employees__status="active"))).order_by("name")
        for ri,s in enumerate(sites,4):
            ws2.cell(row=ri,column=1).value=ri-3
            ws2.cell(row=ri,column=2).value=s.district.state.name
            ws2.cell(row=ri,column=3).value=s.district.name
            ws2.cell(row=ri,column=4).value=s.name
            ws2.cell(row=ri,column=5).value=s.sanctioned_strength
            ws2.cell(row=ri,column=6).value=s.dep
        _apply_table(ws2,3,sites.count()+4,6,"6A0DAD")

        # Sheet 3 — Payroll summary
        ws3=wb.create_sheet("Payroll")
        run=PayrollRun.objects.filter(month=month,year=year).first()
        _title_row(ws3,f"PAYROLL SUMMARY — {mname.upper()} {year}","14","15966A")
        if run:
            agg=run.payslips.aggregate(
                gross=Sum("basic")+Sum("hra")+Sum("da")+Sum("other_allowances"),
                pf=Sum("pf_employee"),esi=Sum("esi_employee"),net=Sum("net_pay"))
            ws3.cell(row=3,column=1).value="Metric"
            ws3.cell(row=3,column=2).value="Amount (INR)"
            ws3.column_dimensions["A"].width=26; ws3.column_dimensions["B"].width=18
            for ri,(k,v) in enumerate([
                ("Total Employees",run.payslips.count()),
                ("Total Gross Wages",float(agg["gross"] or 0)),
                ("Total PF (Employee)",float(agg["pf"] or 0)),
                ("Total ESI (Employee)",float(agg["esi"] or 0)),
                ("Total Net Payable",float(agg["net"] or 0)),
                ("Run Status",run.run_status.upper()),
            ],4):
                ws3.cell(row=ri,column=1).value=k
                ws3.cell(row=ri,column=2).value=round(v,2) if isinstance(v,float) else v
            _apply_table(ws3,3,10,2,"15966A")

        # Sheet 4 — Recruitment snapshot
        ws4=wb.create_sheet("Recruitment")
        _title_row(ws4,"RECRUITMENT SNAPSHOT",3,"E8821E")
        ws4.cell(row=3,column=1).value="Stage"
        ws4.cell(row=3,column=2).value="Count"
        ws4.cell(row=3,column=3).value="%"
        ws4.column_dimensions["A"].width=16; ws4.column_dimensions["B"].width=10; ws4.column_dimensions["C"].width=10
        from apps.recruitment.models import Candidate as RC
        rcounts=dict(RC.objects.values_list("stage").annotate(n=Cnt("id")))
        rtotal=sum(rcounts.values()) or 1
        for ri,(s,label) in enumerate(
            [("applied","Applied"),("screened","Screened"),("interview","Interview"),
             ("selected","Selected"),("rejected","Rejected")],4):
            n=rcounts.get(s,0)
            ws4.cell(row=ri,column=1).value=label
            ws4.cell(row=ri,column=2).value=n
            ws4.cell(row=ri,column=3).value=round(n/rtotal*100,1)
        _apply_table(ws4,3,9,3,"E8821E")

        return _xlsx_response(wb,f"mis_pack_{year}_{month:02d}.xlsx")


# ── Report list ───────────────────────────────────────────────
class ReportListView(APIView):
    permission_classes = [IsAdminHR]

    def get(self, request):
        return Response({
            "reports": [
                {"key":"payroll",     "label":"Payroll Statement",          "url":"/reports/payroll-statement/",    "needs_month":True},
                {"key":"attendance",  "label":"Monthly Attendance Summary", "url":"/reports/attendance-summary/",   "needs_month":True},
                {"key":"deployment",  "label":"Deployment Strength",        "url":"/reports/deployment-strength/",  "needs_month":False},
                {"key":"pf_esi",      "label":"PF/ESI Deduction Register",  "url":"/reports/deduction-register/",   "needs_month":True},
                {"key":"recruitment", "label":"Recruitment Status",         "url":"/reports/recruitment-status/",   "needs_month":False},
                {"key":"mis_pack",    "label":"MIS Pack (All Sheets)",      "url":"/reports/mis-pack/",             "needs_month":True},
            ]
        })
